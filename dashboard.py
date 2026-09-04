import dash
from flask import request, Response
from dash import dcc, html, Input, Output, State, dash_table, MATCH, ALL
import dash_bootstrap_components as dbc
from database import get_connection, init_sqlite_db
from datetime import datetime, date, timedelta
import pandas as pd
import folium
from folium.plugins import MarkerCluster, Fullscreen, MiniMap, HeatMap
import os
import cv2
import base64
from ultralytics import YOLO
import plotly.express as px
import plotly.graph_objects as go
from PIL import Image
from PIL.ExifTags import TAGS, GPSTAGS
from geopy.geocoders import Nominatim
import time
import json
import requests
import io
import math
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

# ----- Configuration Dash -----
app = dash.Dash(
    __name__,
    external_stylesheets=[
        "https://cdnjs.cloudflare.com/ajax/libs/bootstrap/5.3.2/css/bootstrap.min.css",
        "https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.5.1/css/all.min.css"
    ],
    suppress_callback_exceptions=True,
    title="SANUYA - Dashboard"
)

server = app.server
#--- Style ---
app.index_string = '''
<!DOCTYPE html>
<html>
    <head>
        {%metas%}
        <title>{%title%}</title>
        {%favicon%}
        {%css%}
        <link href="https://cdnjs.cloudflare.com/ajax/libs/bootstrap/5.3.2/css/bootstrap.min.css" rel="stylesheet">
        <link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.5.1/css/all.min.css" rel="stylesheet">
        <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap" rel="stylesheet">
        <style>
            * { font-family: 'Inter', system-ui, -apple-system, sans-serif; box-sizing: border-box; }
            body { background: #f8fafc; color: #1e293b; margin: 0; font-size: 14px; }
            
            ::-webkit-scrollbar { width: 6px; height: 6px; }
            ::-webkit-scrollbar-track { background: #f1f5f9; }
            ::-webkit-scrollbar-thumb { background: #cbd5e1; border-radius: 4px; }
            ::-webkit-scrollbar-thumb:hover { background: #94a3b8; }

            .sidebar-link {
                transition: all 0.15s ease-in-out;
                color: #94a3b8 !important;
                font-weight: 500;
                font-size: 13.5px;
                border-left: 3px solid transparent;
            }
            .sidebar-link:hover {
                background: rgba(255,255,255,0.06);
                color: #f8fafc !important;
            }
            .sidebar-link.active {
                background: rgba(56, 189, 248, 0.12);
                color: #ffffff !important;
                font-weight: 600;
                border-left: 3px solid #38bdf8;
            }
            .sidebar-link i { width: 22px; text-align: center; }
            
            .page-title {
                font-weight: 800;
                color: #0f172a;
                font-size: 26px;
                letter-spacing: -0.6px;
                margin-bottom: 2px;
            }
            .page-subtitle {
                color: #64748b;
                font-size: 14px;
                margin-bottom: 0;
            }
            
            .stat-card {
                background: #ffffff;
                border-radius: 12px;
                padding: 18px 20px;
                box-shadow: 0 1px 3px 0 rgba(0,0,0,0.04), 0 1px 2px -1px rgba(0,0,0,0.04);
                border: 1px solid #e2e8f0;
                transition: transform 0.15s ease, box-shadow 0.15s ease;
            }
            .stat-card:hover {
                transform: translateY(-2px);
                box-shadow: 0 6px 20px rgba(0,0,0,0.05);
                border-color: #cbd5e1;
            }

            .content-card {
                background: #ffffff;
                border-radius: 14px;
                padding: 22px 24px;
                box-shadow: 0 1px 3px rgba(0,0,0,0.04);
                border: 1px solid #e2e8f0;
            }

            .map-container {
                height: 620px;
                border-radius: 10px;
                overflow: hidden;
                border: 1px solid #e2e8f0;
                background-color: #f1f5f9;
            }
            .map-container iframe {
                width: 100%;
                height: 100%;
                border: none;
                display: block;
            }

            .legend {
                display: flex;
                gap: 16px;
                padding: 12px 18px;
                background: #f8fafc;
                border-radius: 10px;
                border: 1px solid #e2e8f0;
                margin-top: 14px;
                flex-wrap: wrap;
            }
            .legend-item {
                display: flex;
                align-items: center;
                gap: 8px;
                font-size: 12.5px;
                color: #334155;
            }
            .legend-dot {
                width: 10px;
                height: 10px;
                border-radius: 50%;
                display: inline-block;
            }

            .pulse-dot {
                width: 8px;
                height: 8px;
                background-color: #10b981;
                border-radius: 50%;
                display: inline-block;
                box-shadow: 0 0 0 0 rgba(16, 185, 129, 0.7);
                animation: pulse 1.8s infinite;
            }
            @keyframes pulse {
                0% { transform: scale(0.95); box-shadow: 0 0 0 0 rgba(16, 185, 129, 0.7); }
                70% { transform: scale(1); box-shadow: 0 0 0 6px rgba(16, 185, 129, 0); }
                100% { transform: scale(0.95); box-shadow: 0 0 0 0 rgba(16, 185, 129, 0); }
            }

            .upload-area {
                border: 2px dashed #cbd5e1;
                border-radius: 12px;
                padding: 36px 20px;
                text-align: center;
                cursor: pointer;
                background: #f8fafc;
                transition: all 0.2s ease;
            }
            .upload-area:hover {
                border-color: #0284c7;
                background: #f0f9ff;
            }

            .depot-card {
                background: #ffffff;
                border-radius: 10px;
                padding: 14px 18px;
                margin-bottom: 8px;
                border: 1px solid #e2e8f0;
                transition: all 0.15s ease;
                display: flex;
                justify-content: space-between;
                align-items: center;
                flex-wrap: wrap;
                gap: 12px;
            }
            .depot-card:hover {
                box-shadow: 0 4px 12px rgba(0,0,0,0.04);
                border-color: #94a3b8;
            }
            .depot-card .depot-info {
                display: flex;
                align-items: center;
                gap: 12px;
                flex-wrap: wrap;
            }
            .depot-card .depot-id {
                font-weight: 800;
                font-size: 14px;
                color: #0f172a;
                min-width: 48px;
                padding: 2px 8px;
                background: #f1f5f9;
                border-radius: 6px;
                text-align: center;
            }
            .depot-card .depot-date {
                font-size: 12.5px;
                color: #64748b;
            }
            .depot-card .depot-volume {
                font-weight: 700;
                color: #0f172a;
                font-size: 13px;
                background: #f8fafc;
                padding: 2px 8px;
                border-radius: 6px;
                border: 1px solid #e2e8f0;
            }
            .depot-card .badge {
                padding: 4px 10px;
                border-radius: 20px;
                font-size: 11.5px;
                font-weight: 600;
                letter-spacing: 0.2px;
            }
            .depot-card .depot-actions {
                display: flex;
                gap: 6px;
                flex-wrap: wrap;
            }
            .depot-card .action-btn {
                padding: 5px 12px;
                border: 1px solid transparent;
                border-radius: 6px;
                font-size: 12px;
                font-weight: 600;
                cursor: pointer;
                transition: all 0.15s ease;
                display: inline-flex;
                align-items: center;
            }
            .depot-card .action-btn:hover {
                transform: translateY(-1px);
                box-shadow: 0 2px 5px rgba(0,0,0,0.05);
            }
            .btn-status { background: #fffbeb; color: #b45309; border-color: #fde68a; }
            .btn-status:hover { background: #fef3c7; }
            .btn-priorite { background: #f0fdf4; color: #15803d; border-color: #bbf7d0; }
            .btn-priorite:hover { background: #dcfce7; }
            .btn-delete { background: #fef2f2; color: #b91c1c; border-color: #fecaca; }
            .btn-delete:hover { background: #fee2e2; }
            .btn-photo { background: #eff6ff; color: #1d4ed8; border-color: #bfdbfe; }
            .btn-photo:hover { background: #dbeafe; }
            .btn-maps { background: #f0fdfa; color: #0f766e; border-color: #99f6e4; }
            .btn-maps:hover { background: #ccfbf1; }
            
            .badge-urgent { background: #fef2f2; color: #dc2626; border: 1px solid #fecaca; }
            .badge-moyen { background: #fffbeb; color: #d97706; border: 1px solid #fde68a; }
            .badge-normal { background: #f0fdf4; color: #16a34a; border: 1px solid #bbf7d0; }
            .badge-en_attente, .badge-attente { background: #fff7ed; color: #c2410c; border: 1px solid #ffedd5; }
            .badge-en_cours, .badge-cours { background: #eff6ff; color: #1d4ed8; border: 1px solid #dbeafe; }
            .badge-resolu { background: #f0fdf4; color: #15803d; border: 1px solid #bbf7d0; }

            .modal-content {
                border-radius: 14px;
                border: 1px solid #e2e8f0;
                box-shadow: 0 20px 25px -5px rgba(0, 0, 0, 0.1), 0 8px 10px -6px rgba(0, 0, 0, 0.1);
            }
            .modal-header {
                border-bottom: 1px solid #f1f5f9;
                font-weight: 700;
                padding: 16px 20px;
            }
            .modal-footer {
                border-top: 1px solid #f1f5f9;
                padding: 12px 20px;
            }

            /* Masquer la barre d'outils Dash DevTools, le bouton violet << et promotion Plotly Cloud */
            [class*="dash-debug"], 
            .dash-debug-menu, 
            .dash-debug-menu__outer, 
            .dash-debug-menu__toggle, 
            .dash-debug-alert, 
            ._dash-devtools, 
            .dash-fe-error__overlay {
                display: none !important;
                visibility: hidden !important;
                opacity: 0 !important;
                pointer-events: none !important;
            }
        </style>
    </head>
    <body>
        {%app_entry%}
        <footer>
            {%config%}
            {%scripts%}
            {%renderer%}
        </footer>
    </body>
</html>
'''

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# --- Chargement du modèle YOLO ---
def load_model():
    paths = [
        os.path.join(BASE_DIR, "runs", "detect", "train", "weights", "best.pt"),
        os.path.join(BASE_DIR, "runs", "detect", "train-2", "weights", "best.pt"),
        os.path.join(BASE_DIR, "yolov8m.pt"),
        os.path.join(BASE_DIR, "yolov8n.pt")
    ]
    for p in paths:
        if os.path.exists(p):
            print(f"[OK] Modèle chargé : {p}")
            return YOLO(p)
    print("[AVERTISSEMENT] Modèle par défaut yolov8n.pt")
    return YOLO('yolov8n.pt')

model = load_model()

# --- Modules Métier SANUYA ---
from database import get_connection, init_sqlite_db
from estimation import estimer_volume, prioriser
from verification import est_doublon

init_sqlite_db()

# --- Fonctions d'extraction de métadonnées ---
def get_location_from_metadata(chemin_photo):
    try:
        image = Image.open(chemin_photo)
        exifdata = image.getexif()
        if not exifdata:
            return None
        
        location_info = []
        
        for tag_id, value in exifdata.items():
            tag = TAGS.get(tag_id, tag_id)
            
            if tag in ['XPComment', 'XPSubject', 'XPKeywords', 'XPTitle', 'ImageDescription', 'Make', 'Model']:
                if isinstance(value, bytes):
                    try:
                        for encoding in ['utf-16le', 'utf-8', 'latin1']:
                            try:
                                decoded = value.decode(encoding, errors='ignore').strip('\x00')
                                if decoded and len(decoded) > 2 and not decoded.startswith('Android'):
                                    location_info.append(decoded)
                                    break
                            except:
                                continue
                    except:
                        pass
                elif isinstance(value, str) and len(value) > 2:
                    location_info.append(value)
        
        if location_info:
            for info in location_info:
                if info and len(info) > 2 and not any(x in info.lower() for x in ['android', 'iphone', 'camera']):
                    return f"Lieu détecté: {info}"
        return None
    except Exception as e:
        print(f"⚠️ Erreur extraction lieu: {e}")
        return None

def get_gps_from_filename(chemin_photo):
    filename = os.path.basename(chemin_photo)
    if 'IMG' in filename and '_' in filename:
        parts = filename.split('_')
        if len(parts) >= 2:
            date_part = parts[1]
            if len(date_part) >= 14:
                try:
                    year = int(date_part[0:4])
                    month = int(date_part[4:6])
                    day = int(date_part[6:8])
                    hour = int(date_part[9:11]) if len(date_part) > 11 else 0
                    minute = int(date_part[11:13]) if len(date_part) > 13 else 0
                    return f"Prise le {day:02d}/{month:02d}/{year} à {hour:02d}h{minute:02d}"
                except:
                    pass
    return None

def get_all_exif_tags(chemin_photo):
    try:
        image = Image.open(chemin_photo)
        exifdata = image.getexif()
        if not exifdata:
            return {}
        
        tags = {}
        for tag_id, value in exifdata.items():
            tag = TAGS.get(tag_id, tag_id)
            if isinstance(value, bytes):
                try:
                    value = value.decode('utf-8', errors='ignore')
                except:
                    value = str(value)
            tags[tag] = value
        return tags
    except Exception as e:
        print(f"⚠️ Erreur tags EXIF: {e}")
        return {}

def search_coordinates_by_name(location_name):
    if not location_name or len(location_name) < 3:
        return None
    
    try:
        geolocator = Nominatim(user_agent="sanuya_app", timeout=5)
        location_name = location_name.replace('Lieu détecté: ', '').strip()
        
        search_terms = [
            location_name,
            f"{location_name}, Mali",
            f"{location_name}, Bamako, Mali"
        ]
        
        for term in search_terms:
            try:
                location = geolocator.geocode(term, language='fr')
                if location:
                    return {
                        'latitude': location.latitude,
                        'longitude': location.longitude,
                        'address': location.address
                    }
            except:
                continue
        
        return None
    except Exception as e:
        print(f"⚠️ Erreur recherche coordonnées: {e}")
        return None

# --- Récuperation du gps ---
def get_gps_from_exif(chemin_photo):
    try:
        image = Image.open(chemin_photo)
        exifdata = image.getexif()
        if not exifdata:
            return None
        gps_info = {}
        for tag_id, value in exifdata.items():
            tag = TAGS.get(tag_id, tag_id)
            if tag == "GPSInfo":
                for gps_tag_id, gps_value in value.items():
                    gps_tag = GPSTAGS.get(gps_tag_id, gps_tag_id)
                    gps_info[gps_tag] = gps_value
        if not gps_info:
            return None
        def convert_to_degrees(value):
            d = float(value[0])
            m = float(value[1])
            s = float(value[2])
            return d + (m / 60.0) + (s / 3600.0)
        lat = convert_to_degrees(gps_info['GPSLatitude'])
        lon = convert_to_degrees(gps_info['GPSLongitude'])
        if gps_info['GPSLatitudeRef'] == 'S':
            lat = -lat
        if gps_info['GPSLongitudeRef'] == 'W':
            lon = -lon
        return {'latitude': lat, 'longitude': lon}
    except Exception as e:
        print(f"⚠️ Erreur lecture GPS : {e}")
        return None

def get_gps_precision(chemin_photo):
    try:
        image = Image.open(chemin_photo)
        exifdata = image.getexif()
        if not exifdata:
            return None
        gps_info = {}
        for tag_id, value in exifdata.items():
            tag = TAGS.get(tag_id, tag_id)
            if tag == "GPSInfo":
                for gps_tag_id, gps_value in value.items():
                    gps_tag = GPSTAGS.get(gps_tag_id, gps_tag_id)
                    gps_info[gps_tag] = gps_value
        if not gps_info:
            return None
        precision = gps_info.get('GPSDOP', None)
        if precision:
            if isinstance(precision, tuple) and len(precision) >= 2:
                return float(precision[0]) / float(precision[1])
            elif isinstance(precision, (int, float)):
                return float(precision)
        return None
    except Exception as e:
        print(f"⚠️ Erreur précision GPS : {e}")
        return None

def get_capture_date(chemin_photo):
    try:
        image = Image.open(chemin_photo)
        exifdata = image.getexif()
        if not exifdata:
            return datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        for tag_id, value in exifdata.items():
            tag = TAGS.get(tag_id, tag_id)
            if tag == "DateTimeOriginal":
                date_str = value
                try:
                    dt = datetime.strptime(date_str, "%Y:%m:%d %H:%M:%S")
                    return dt.strftime("%d/%m/%Y %H:%M:%S")
                except:
                    return date_str
        return datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    except Exception as e:
        print(f"⚠️ Erreur date capture : {e}")
        return datetime.now().strftime("%d/%m/%Y %H:%M:%S")

# --- Géolocalisation inverse & Moteur de Quartiers de Bamako ---
geolocator = Nominatim(user_agent="sanuya_dashboard_app")
_address_cache = {}

# Référentiel géographique exhaustif des 65+ quartiers officiels de Bamako par Commune
QUARTIERS_BAMAKO = [
    # Commune I (Rive Gauche Nord-Est)
    {"nom": "Banconi", "commune": "Commune I", "lat": 12.6820, "lon": -7.9730},
    {"nom": "Boulkassoumbougou", "commune": "Commune I", "lat": 12.6710, "lon": -7.9520},
    {"nom": "Djelibougou", "commune": "Commune I", "lat": 12.6560, "lon": -7.9510},
    {"nom": "Doumanzana", "commune": "Commune I", "lat": 12.6860, "lon": -7.9340},
    {"nom": "Fadjiguila", "commune": "Commune I", "lat": 12.6660, "lon": -7.9290},
    {"nom": "Korofina Nord", "commune": "Commune I", "lat": 12.6620, "lon": -7.9640},
    {"nom": "Korofina Sud", "commune": "Commune I", "lat": 12.6490, "lon": -7.9660},
    {"nom": "Sikoroni", "commune": "Commune I", "lat": 12.6970, "lon": -7.9710},
    {"nom": "Sotuba", "commune": "Commune I", "lat": 12.6460, "lon": -7.9170},
    
    # Commune II (Rive Gauche Centre-Est)
    {"nom": "Bakaribougou", "commune": "Commune II", "lat": 12.6650, "lon": -7.9810},
    {"nom": "Bougouba", "commune": "Commune II", "lat": 12.6580, "lon": -7.9420},
    {"nom": "Bozola", "commune": "Commune II", "lat": 12.6410, "lon": -7.9860},
    {"nom": "Hippodrome", "commune": "Commune II", "lat": 12.6610, "lon": -7.9860},
    {"nom": "Hippodrome II", "commune": "Commune II", "lat": 12.6690, "lon": -7.9870},
    {"nom": "Médina-Coura", "commune": "Commune II", "lat": 12.6480, "lon": -7.9910},
    {"nom": "Missira", "commune": "Commune II", "lat": 12.6510, "lon": -7.9810},
    {"nom": "Niaréla", "commune": "Commune II", "lat": 12.6430, "lon": -7.9790},
    {"nom": "Quinzambougou", "commune": "Commune II", "lat": 12.6440, "lon": -7.9870},
    {"nom": "TSF-Sans Fil", "commune": "Commune II", "lat": 12.6530, "lon": -7.9720},
    {"nom": "Zone Industrielle", "commune": "Commune II", "lat": 12.6420, "lon": -7.9620},
    
    # Commune III (Rive Gauche Centre-Ouest / Colline)
    {"nom": "Bolibana", "commune": "Commune III", "lat": 12.6392, "lon": -8.0035},
    {"nom": "Darsalam", "commune": "Commune III", "lat": 12.6460, "lon": -8.0060},
    {"nom": "Centre Commercial", "commune": "Commune III", "lat": 12.6370, "lon": -7.9970},
    {"nom": "Dravela", "commune": "Commune III", "lat": 12.6320, "lon": -8.0090},
    {"nom": "Ouolofobougou", "commune": "Commune III", "lat": 12.6360, "lon": -8.0120},
    {"nom": "Badialan I", "commune": "Commune III", "lat": 12.6470, "lon": -8.0130},
    {"nom": "Badialan II", "commune": "Commune III", "lat": 12.6510, "lon": -8.0160},
    {"nom": "Badialan III", "commune": "Commune III", "lat": 12.6550, "lon": -8.0190},
    {"nom": "N'Tomikorobougou", "commune": "Commune III", "lat": 12.6430, "lon": -8.0170},
    {"nom": "Bamako-Coura", "commune": "Commune III", "lat": 12.6330, "lon": -8.0030},
    {"nom": "Niomirambougou", "commune": "Commune III", "lat": 12.6520, "lon": -8.0210},
    {"nom": "Koulouba", "commune": "Commune III", "lat": 12.6690, "lon": -8.0160},
    {"nom": "Point G", "commune": "Commune III", "lat": 12.6760, "lon": -7.9960},
    {"nom": "Samé", "commune": "Commune III", "lat": 12.6710, "lon": -8.0360},
    {"nom": "Sirakoro Dounfing", "commune": "Commune III", "lat": 12.6850, "lon": -8.0450},
    
    # Commune IV (Rive Gauche Ouest)
    {"nom": "Hamdallaye", "commune": "Commune IV", "lat": 12.6510, "lon": -8.0210},
    {"nom": "ACI 2000", "commune": "Commune IV", "lat": 12.6330, "lon": -8.0260},
    {"nom": "Lafiabougou", "commune": "Commune IV", "lat": 12.6410, "lon": -8.0360},
    {"nom": "Djicoroni-Para", "commune": "Commune IV", "lat": 12.6210, "lon": -8.0360},
    {"nom": "Sébénikoro", "commune": "Commune IV", "lat": 12.6060, "lon": -8.0660},
    {"nom": "Taliko", "commune": "Commune IV", "lat": 12.6260, "lon": -8.0560},
    {"nom": "Lassa", "commune": "Commune IV", "lat": 12.6560, "lon": -8.0510},
    {"nom": "Sibiribougou", "commune": "Commune IV", "lat": 12.5960, "lon": -8.0860},
    {"nom": "Kalabambougou", "commune": "Commune IV", "lat": 12.5860, "lon": -8.0510},
    
    # Commune V (Rive Droite Ouest-Centre)
    {"nom": "Badalabougou", "commune": "Commune V", "lat": 12.6240, "lon": -7.9860},
    {"nom": "Quartier Mali", "commune": "Commune V", "lat": 12.6190, "lon": -7.9960},
    {"nom": "Torokorobougou", "commune": "Commune V", "lat": 12.6110, "lon": -7.9910},
    {"nom": "Baco-Djicoroni", "commune": "Commune V", "lat": 12.5860, "lon": -8.0160},
    {"nom": "Sabalibougou", "commune": "Commune V", "lat": 12.6060, "lon": -7.9760},
    {"nom": "Daoudabougou", "commune": "Commune V", "lat": 12.6010, "lon": -7.9560},
    {"nom": "Kalaban-Coura", "commune": "Commune V", "lat": 12.5860, "lon": -7.9760},
    {"nom": "Sema I", "commune": "Commune V", "lat": 12.6160, "lon": -7.9810},
    
    # Commune VI (Rive Droite Est)
    {"nom": "Sogoniko", "commune": "Commune VI", "lat": 12.6010, "lon": -7.9410},
    {"nom": "Faladié", "commune": "Commune VI", "lat": 12.5860, "lon": -7.9460},
    {"nom": "Banankabougou", "commune": "Commune VI", "lat": 12.5960, "lon": -7.9260},
    {"nom": "Magnambougou", "commune": "Commune VI", "lat": 12.6110, "lon": -7.9360},
    {"nom": "Niamakoro", "commune": "Commune VI", "lat": 12.5860, "lon": -7.9660},
    {"nom": "Yirimadio", "commune": "Commune VI", "lat": 12.5960, "lon": -7.8960},
    {"nom": "Missabougou", "commune": "Commune VI", "lat": 12.6260, "lon": -7.9160},
    {"nom": "Sokorodji", "commune": "Commune VI", "lat": 12.5760, "lon": -7.9560},
    {"nom": "Sénou", "commune": "Commune VI", "lat": 12.5360, "lon": -7.9310},
    {"nom": "Djanékéla", "commune": "Commune VI", "lat": 12.5710, "lon": -7.9110},
]

def find_nearest_quartier(lat, lon):
    best_q = None
    min_dist = float('inf')
    for q in QUARTIERS_BAMAKO:
        d_lat = (lat - q['lat']) * 111.0
        d_lon = (lon - q['lon']) * 111.0 * math.cos(math.radians(lat))
        dist_km = math.sqrt(d_lat**2 + d_lon**2)
        if dist_km < min_dist:
            min_dist = dist_km
            best_q = q
    return best_q, min_dist * 1000.0

def get_location_details(lat, lon):
    try:
        lat_f, lon_f = float(lat), float(lon)
    except (ValueError, TypeError):
        return {
            'quartier': 'Inconnu',
            'commune': 'Bamako',
            'repere': '',
            'adresse_complete': 'Coordonnées inconnues'
        }

    key = (round(lat_f, 4), round(lon_f, 4))
    if key in _address_cache:
        return _address_cache[key]

    nearest_q, dist_m = find_nearest_quartier(lat_f, lon_f)
    quartier_nom = nearest_q['nom'] if nearest_q else 'Bamako'
    commune_nom = nearest_q['commune'] if nearest_q else get_commune_bamako(lat_f, lon_f)
    repere_ou_rue = ""

    try:
        location = geolocator.reverse((lat_f, lon_f), timeout=3, language='fr')
        if location and location.raw:
            raw_addr = location.raw.get('address', {})
            osm_q = (raw_addr.get('quarter') or 
                     raw_addr.get('neighbourhood') or 
                     raw_addr.get('suburb') or 
                     raw_addr.get('residential'))
            
            if osm_q and osm_q.lower() not in ('centre ville', 'bamako', 'district de bamako', 'mali'):
                quartier_nom = osm_q
                
            rep = raw_addr.get('road') or raw_addr.get('amenity') or raw_addr.get('building')
            if rep:
                repere_ou_rue = rep
    except Exception:
        pass

    adresse_formatee = f"Quartier {quartier_nom}"
    if repere_ou_rue:
        adresse_formatee += f" ({repere_ou_rue})"

    res = {
        'quartier': quartier_nom,
        'commune': commune_nom,
        'repere': repere_ou_rue,
        'adresse_complete': adresse_formatee
    }
    _address_cache[key] = res
    return res

def get_address(lat, lon):
    info = get_location_details(lat, lon)
    return info['adresse_complete']

# ==================== DÉCOUPAGE DES 6 COMMUNES DE BAMAKO ====================
COMMUNES_BAMAKO = {
    'Commune I': {
        'nom': 'Commune I',
        'description': 'Rive gauche Nord-Est • Djelibougou, Korofina, Sotuba, Doumanzana, Fadjiguila',
        'couleur': '#2563eb',
        'polygone': [
            [12.640, -7.970], [12.670, -7.965], [12.710, -7.935], 
            [12.675, -7.900], [12.635, -7.935], [12.640, -7.970]
        ]
    },
    'Commune II': {
        'nom': 'Commune II',
        'description': 'Rive gauche Centre-Est • Médina-Coura, Bozola, Hippodrome, Missira, Niaréla',
        'couleur': '#7c3aed',
        'polygone': [
            [12.628, -8.010], [12.642, -7.995], [12.680, -7.990], 
            [12.670, -7.965], [12.640, -7.970], [12.632, -7.995], [12.628, -8.010]
        ]
    },
    'Commune III': {
        'nom': 'Commune III',
        'description': 'Rive gauche Centre-Ouest • Centre commercial, Darsalam, Badialan, Koulouba, Point-G',
        'couleur': '#db2777',
        'polygone': [
            [12.625, -8.025], [12.650, -8.025], [12.700, -8.015], 
            [12.680, -7.990], [12.642, -7.995], [12.628, -8.010], [12.625, -8.025]
        ]
    },
    'Commune IV': {
        'nom': 'Commune IV',
        'description': 'Rive gauche Ouest • Hamdallaye, Lafiabougou, Djicoroni-Para, Sebenikoro, Sibiribougou',
        'couleur': '#d97706',
        'polygone': [
            [12.580, -8.100], [12.640, -8.070], [12.650, -8.025], 
            [12.625, -8.025], [12.590, -8.055], [12.580, -8.100]
        ]
    },
    'Commune V': {
        'nom': 'Commune V',
        'description': 'Rive droite Centre • Badalabougou, Sema I, Torokorobougou, Daoudabougou, Sabalibougou',
        'couleur': '#059669',
        'polygone': [
            [12.590, -8.055], [12.625, -8.010], [12.630, -7.995], 
            [12.620, -7.970], [12.585, -7.975], [12.555, -8.020], 
            [12.570, -8.050], [12.590, -8.055]
        ]
    },
    'Commune VI': {
        'nom': 'Commune VI',
        'description': 'Rive droite Sud-Est • Sogoniko, Magnambougou, Banankabougou, Faladié, Niamakoro, Yirimadio',
        'couleur': '#0891b2',
        'polygone': [
            [12.620, -7.970], [12.635, -7.935], [12.620, -7.900], 
            [12.545, -7.925], [12.540, -7.975], [12.585, -7.975], [12.620, -7.970]
        ]
    }
}

def point_in_polygon(lat, lon, polygon):
    n = len(polygon)
    inside = False
    p1x, p1y = polygon[0]
    for i in range(1, n + 1):
        p2x, p2y = polygon[i % n]
        if lat > min(p1x, p2x):
            if lat <= max(p1x, p2x):
                if lon <= max(p1y, p2y):
                    if p1x != p2x:
                        xinters = (lat - p1x) * (p2y - p1y) / (p2x - p1x) + p1y
                    if p1y == p2y or lon <= xinters:
                        inside = not inside
        p1x, p1y = p2x, p2y
    return inside

def get_commune_bamako(lat, lon):
    try:
        lat_f, lon_f = float(lat), float(lon)
    except:
        return 'Bamako (Général)'
    for c_id, c_data in COMMUNES_BAMAKO.items():
        if point_in_polygon(lat_f, lon_f, c_data['polygone']):
            return c_data['nom']
    # Fallback par proximité du centre de la commune
    min_dist = float('inf')
    closest = 'Bamako (Général)'
    for c_id, c_data in COMMUNES_BAMAKO.items():
        clat, clon = c_data['polygone'][0]
        d = (lat_f - clat)**2 + (lon_f - clon)**2
        if d < min_dist:
            min_dist = d
            closest = c_data['nom']
    return closest

# ==================== CLASSIFICATION MULTI-PRODUITS & DÉCHETS ====================
CATEGORIES_PRODUITS = {
    'bottle': ('Plastique & Bouteille', '#0284c7', (199, 132, 2)),
    'cup': ('Gobelet & Emballage', '#06b6d4', (212, 182, 6)),
    'can': ('Métal & Canette', '#eab308', (8, 179, 234)),
    'bowl': ('Récipient plastique', '#0284c7', (199, 132, 2)),
    'box': ('Carton & Emballage', '#d97706', (6, 119, 217)),
    'bag': ('Sachet plastique', '#6366f1', (241, 102, 99)),
    'backpack': ('Textile & Sac', '#8b5cf6', (246, 92, 139)),
    'suitcase': ('Déchet volumineux', '#ef4444', (68, 68, 239)),
    'chair': ('Mobilier / Encombrant', '#f97316', (22, 115, 249)),
    'bench': ('Mobilier urbain', '#f97316', (22, 115, 249)),
}

def categoriser_produit(label):
    lbl = str(label).lower().strip()
    if lbl in CATEGORIES_PRODUITS:
        return CATEGORIES_PRODUITS[lbl]
    if any(k in lbl for k in ['plastic', 'bottle', 'sac', 'poubelle', 'bag']):
        return ('Plastique & Sachet', '#0284c7', (199, 132, 2))
    if any(k in lbl for k in ['carton', 'paper', 'box', 'emballage']):
        return ('Carton & Papier', '#d97706', (6, 119, 217))
    if any(k in lbl for k in ['metal', 'can', 'fer', 'canette']):
        return ('Métal & Ferraille', '#eab308', (8, 179, 234))
    if any(k in lbl for k in ['pneu', 'tire', 'wheel', 'caoutchouc']):
        return ('Pneu & Caoutchouc', '#8b5cf6', (246, 92, 139))
    if any(k in lbl for k in ['gravat', 'stone', 'rock', 'brick', 'debris']):
        return ('Gravats & Décombres', '#ef4444', (68, 68, 239))
    return ('Ordures ménagères mixtes', '#10b981', (129, 185, 16))

# ==================== STATS POUR LE TABLEAU DE BORD ====================
def get_stats_dashboard():
    conn = get_connection()
    if not conn:
        return {}
    try:
        cursor = conn.cursor()
        
        cursor.execute("SELECT COUNT(*) FROM signalements WHERE statut != 'resolu'")
        total = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM signalements WHERE priorite = 'urgent' AND statut != 'resolu'")
        urgent = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM signalements WHERE priorite = 'moyen' AND statut != 'resolu'")
        moyen = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM signalements WHERE priorite = 'normal' AND statut != 'resolu'")
        normal = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM signalements WHERE statut = 'en_attente'")
        attente = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM signalements WHERE statut = 'en_cours'")
        cours = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM signalements WHERE statut = 'resolu'")
        resolu = cursor.fetchone()[0]

        cursor.execute("SELECT COALESCE(SUM(volume), 0) FROM signalements WHERE statut != 'resolu'")
        vol_row = cursor.fetchone()
        vol_total = float(vol_row[0]) if vol_row and vol_row[0] is not None else 0.0

        conn.close()
        taux_res = round((resolu / (total + resolu)) * 100, 1) if (total + resolu) > 0 else 0.0
        return {
            'total': total, 'urgent': urgent, 'moyen': moyen, 'normal': normal,
            'attente': attente, 'cours': cours, 'resolu': resolu,
            'volume_total': round(vol_total, 2),
            'taux_resolution': taux_res
        }
    except Exception as e:
        print(f"❌ Erreur stats dashboard : {e}")
        return {}

def get_stats_complete():
    conn = get_connection()
    if not conn:
        return {}
    try:
        cursor = conn.cursor()
        
        cursor.execute("SELECT COUNT(*) FROM signalements")
        total = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM signalements WHERE priorite = 'urgent'")
        urgent = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM signalements WHERE priorite = 'moyen'")
        moyen = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM signalements WHERE priorite = 'normal'")
        normal = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM signalements WHERE statut = 'en_attente'")
        attente = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM signalements WHERE statut = 'en_cours'")
        cours = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM signalements WHERE statut = 'resolu'")
        resolu = cursor.fetchone()[0]
        
        conn.close()
        return {'total': total, 'urgent': urgent, 'moyen': moyen, 'normal': normal,
                'attente': attente, 'cours': cours, 'resolu': resolu}
    except Exception as e:
        print(f"❌ Erreur stats complètes : {e}")
        return {}

def get_depots_filtres(filtre_priorite='tous', filtre_statut='tous', filtre_commune='tous'):
    conn = get_connection()
    if not conn:
        return []
    try:
        cursor = conn.cursor(dictionary=True)
        
        query = """
            SELECT id, latitude, longitude, volume, priorite, statut, 
                   DATE_FORMAT(date_creation, '%d/%m/%Y %H:%i') as date,
                   photo_chemin
            FROM signalements 
            WHERE 1=1
        """
        params = []
        
        if filtre_statut == 'actifs':
            query += " AND statut != 'resolu'"
        elif filtre_statut != 'tous':
            query += " AND statut = %s"
            params.append(filtre_statut)
        
        if filtre_priorite != 'tous':
            query += " AND priorite = %s"
            params.append(filtre_priorite)
        
        query += " ORDER BY date_creation DESC"
        
        cursor.execute(query, params)
        data = cursor.fetchall()
        conn.close()
        
        # Attribution automatique de la commune officielle
        for d in data:
            d['commune'] = get_commune_bamako(d.get('latitude'), d.get('longitude'))
            
        if filtre_commune != 'tous':
            data = [d for d in data if d['commune'] == filtre_commune]
            
        return data
    except Exception as e:
        print(f"❌ Erreur depots : {e}")
        return []

# ==================== FONCTIONS CRUD ====================
def add_detection(lat, lon, volume, priorite, statut, photo_nom, photo_chemin=None):
    conn = get_connection()
    if not conn:
        return {'success': False, 'is_doublon': False}
    try:
        if photo_chemin is None:
            photo_chemin = f"images_test/{photo_nom}"
        
        # Vérification anti-doublons par GPS (< 50 mètres)
        existants = get_depots_filtres()
        is_dup, id_dup, distance = est_doublon(lat, lon, existants, seuil=50)
        
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO signalements 
            (latitude, longitude, volume, priorite, statut, date_creation, photo_nom, photo_chemin, est_doublon, doublon_de)
            VALUES (%s, %s, %s, %s, %s, NOW(), %s, %s, %s, %s)
        """, (lat, lon, volume, priorite, statut, photo_nom, photo_chemin, 1 if is_dup else 0, id_dup))
        conn.commit()
        conn.close()
        return {'success': True, 'is_doublon': is_dup, 'id_doublon': id_dup, 'distance': distance}
    except Exception as e:
        print(f"[ERREUR] Erreur ajout : {e}")
        return {'success': False, 'is_doublon': False, 'erreur': str(e)}

def update_status_db(depot_id, statut):
    conn = get_connection()
    if not conn:
        return False
    try:
        cursor = conn.cursor()
        cursor.execute("UPDATE signalements SET statut = %s WHERE id = %s", (statut, depot_id))
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        print(f"❌ Erreur mise à jour : {e}")
        return False

def update_priorite_db(depot_id, priorite):
    conn = get_connection()
    if not conn:
        return False
    try:
        cursor = conn.cursor()
        cursor.execute("UPDATE signalements SET priorite = %s WHERE id = %s", (priorite, depot_id))
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        print(f"❌ Erreur mise à jour : {e}")
        return False

def delete_depot_db(depot_id):
    conn = get_connection()
    if not conn:
        return False
    try:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM signalements WHERE id = %s", (depot_id,))
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        print(f"❌ Erreur suppression : {e}")
        return False

# ==================== ANALYSE YOLO MULTI-PRODUITS AVEC TRACÉ PRÉCIS ====================
def analyser_photo(chemin):
    if not os.path.exists(chemin):
        return {'erreur': 'Fichier introuvable'}
    img = cv2.imread(chemin)
    if img is None:
        return {'erreur': 'Image invalide'}
    
    img_h, img_w = img.shape[:2]
    r = model.predict(img, conf=0.25, verbose=False)
    
    annotated_frame = img.copy()
    overlay = img.copy()
    
    dechets = []
    produits_counts = {}
    
    if len(r[0].boxes) > 0:
        # Étape 1 : Remplissage des masques semi-transparents par produit
        for box in r[0].boxes:
            nom_raw = model.names[int(box.cls[0])]
            conf = float(box.conf[0])
            x1, y1, x2, y2 = map(int, box.xyxy[0].tolist())
            
            cat_nom, cat_hex, cat_bgr = categoriser_produit(nom_raw)
            produits_counts[cat_nom] = produits_counts.get(cat_nom, 0) + 1
            
            # Masque semi-transparent
            cv2.rectangle(overlay, (x1, y1), (x2, y2), cat_bgr, -1)
            
            dechets.append({
                'type': cat_nom,
                'type_raw': nom_raw,
                'confiance': round(conf * 100, 1),
                'couleur_hex': cat_hex,
                'couleur_bgr': cat_bgr,
                'coords': (x1, y1, x2, y2),
                'largeur': x2 - x1,
                'hauteur': y2 - y1
            })
            
        # Fusion par transparence (30% d'opacité)
        cv2.addWeighted(overlay, 0.30, annotated_frame, 0.70, 0, annotated_frame)
        
        # Étape 2 : Tracé précis des contours nets et étiquettes badge
        for d in dechets:
            x1, y1, x2, y2 = d['coords']
            cat_bgr = d['couleur_bgr']
            
            # Bordure externe nette
            cv2.rectangle(annotated_frame, (x1, y1), (x2, y2), cat_bgr, 2)
            
            # Étiquette moderne avec contraste
            tag = f"{d['type']} {d['confiance']:.0f}%"
            (tw, th), _ = cv2.getTextSize(tag, cv2.FONT_HERSHEY_SIMPLEX, 0.42, 1)
            cv2.rectangle(annotated_frame, (x1, max(0, y1 - 20)), (x1 + tw + 8, max(20, y1)), cat_bgr, -1)
            cv2.putText(annotated_frame, tag, (x1 + 4, max(14, y1 - 5)), cv2.FONT_HERSHEY_SIMPLEX, 0.42, (255, 255, 255), 1, cv2.LINE_AA)
            
    nom_base = os.path.basename(chemin)
    chemin_annote = os.path.join(os.path.dirname(chemin), f"annote_{nom_base}")
    cv2.imwrite(chemin_annote, annotated_frame)
    
    # Encodage base64 pour affichage direct
    _, buffer = cv2.imencode('.jpg', annotated_frame)
    img_b64 = base64.b64encode(buffer).decode('utf-8')
    
    if dechets:
        surface_relative = sum(d['largeur'] * d['hauteur'] for d in dechets) / (img_w * img_h)
        vol = max(0.1, round(surface_relative * 15.0, 2))
        prio = prioriser(vol)
        return {
            'nb': len(dechets),
            'dechets': dechets,
            'produits_counts': produits_counts,
            'volume': vol,
            'priorite': prio,
            'image_b64': img_b64,
            'chemin_annote': chemin_annote
        }
    
    return {
        'nb': 0,
        'dechets': [],
        'produits_counts': {},
        'volume': 0,
        'priorite': 'normal',
        'image_b64': img_b64,
        'chemin_annote': chemin_annote
    }

# ==================== GENERATION DE CARTE SIG PRO ====================
def generate_map():
    depots = get_depots_filtres('tous', 'tous', 'tous')
    
    carte = folium.Map(
        location=[12.6392, -8.0029],
        zoom_start=13,
        tiles=None
    )
    
    # Couches de fond cartographique
    folium.TileLayer('OpenStreetMap', name='🗺️ Plan Standard (OSM)', control=True).add_to(carte)
    folium.TileLayer(
        tiles='https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}',
        attr='Esri, Maxar, Earthstar Geographics',
        name='🛰️ Vue Satellite (Esri)',
        control=True
    ).add_to(carte)
    folium.TileLayer(
        tiles='https://{s}.tile.opentopomap.org/{z}/{x}/{y}.png',
        attr='OpenTopoMap',
        name='🏔️ Relief Topographique',
        control=True
    ).add_to(carte)
    
    # Plugins Folium Enterprise
    Fullscreen(position='topright', title='Plein écran', title_cancel='Quitter plein écran', force_separate_button=True).add_to(carte)
    MiniMap(toggle_display=True, position='bottomleft', width=130, height=95, zoom_animation=True).add_to(carte)
    
    # Couche SIG officielle : Découpage des 6 Communes de Bamako
    communes_group = folium.FeatureGroup(name="🏛️ Découpage des 6 Communes (Mairie)", show=True)
    for c_id, c_data in COMMUNES_BAMAKO.items():
        folium.Polygon(
            locations=c_data["polygone"],
            color=c_data["couleur"],
            weight=2,
            dash_array='4, 6',
            fill=True,
            fill_color=c_data["couleur"],
            fill_opacity=0.07,
            tooltip=f"<b>{c_data['nom']}</b><br>{c_data['description']}",
            popup=folium.Popup(f"""
                <div style="font-family: 'Inter', system-ui, sans-serif; min-width: 200px;">
                    <h6 style="margin:0 0 4px 0; font-weight:800; color:{c_data['couleur']};">{c_data['nom']}</h6>
                    <p style="font-size:12px; color:#475569; margin:0; line-height:1.4;">{c_data['description']}</p>
                </div>
            """, max_width=260)
        ).add_to(communes_group)
    communes_group.add_to(carte)
    
    if not depots:
        folium.Marker(
            location=[12.6392, -8.0029],
            popup="Aucun dépôt enregistré",
            icon=folium.Icon(color='gray', icon='info', prefix='fa')
        ).add_to(carte)
        folium.LayerControl(position='topright', collapsed=True).add_to(carte)
        return carte._repr_html_()
    
    # Calques SIG : Les rayons sont ajoutés en fond (non-interactifs) pour laisser les marqueurs au premier plan
    couche_rayons = folium.FeatureGroup(name="⭕ Périmètres d'impact & Zones assainies", show=True).add_to(carte)
    couche_actifs = folium.FeatureGroup(name="🚨 Dépôts Actifs (À évacuer)", show=True).add_to(carte)
    couche_resolus = folium.FeatureGroup(name="✅ Dépôts Traités & Résolus", show=True).add_to(carte)
    
    couleurs = {'urgent': 'red', 'moyen': 'orange', 'normal': 'green'}
    badge_colors = {
        'urgent': ('#fef2f2', '#dc2626'),
        'moyen': ('#fffbeb', '#d97706'),
        'normal': ('#f0fdf4', '#16a34a')
    }
    statut_labels = {'en_attente': 'En attente', 'en_cours': 'En cours', 'resolu': 'Résolu'}
    
    heat_points = []
    
    for depot in depots:
        lat = float(depot['latitude'])
        lon = float(depot['longitude'])
        
        priorite = depot.get('priorite', 'normal')
        statut = depot.get('statut', 'en_attente')
        is_resolu = (statut == 'resolu')
        
        commune_nom = depot.get('commune') or get_commune_bamako(lat, lon)
        try:
            volume = float(depot.get('volume', 0.0))
        except (ValueError, TypeError):
            volume = 0.0
            
        depot_id = depot.get('id', '?')
        photo_path = depot.get('photo_chemin', '')
        
        is_default_coords = (abs(lat - 12.6392) < 0.0005 and abs(lon - (-8.0029)) < 0.0005)
        
        loc_info = get_location_details(lat, lon)
        commune_nom = depot.get('commune') or loc_info['commune']
        quartier_nom = loc_info['quartier']
        adresse = loc_info['adresse_complete']
        statut_text = statut_labels.get(statut, str(statut).capitalize())
        bennes_estimees = max(1, round(volume / 5.0))
        
        photo_thumb = ""
        photo_mini = ""
        actual_photo = photo_path
        if actual_photo and not os.path.isabs(actual_photo):
            actual_photo = os.path.join(os.path.dirname(os.path.abspath(__file__)), actual_photo)
            
        if actual_photo and os.path.exists(actual_photo):
            try:
                with open(actual_photo, "rb") as f:
                    b64 = base64.b64encode(f.read()).decode()
                photo_thumb = f"<div style='margin-bottom:8px; text-align:center;'><img src='data:image/jpeg;base64,{b64}' style='width:100%; max-height:115px; object-fit:cover; border-radius:6px; border:1px solid #e2e8f0;'/></div>"
                photo_mini = f"<div style='margin-bottom:6px; text-align:center;'><img src='data:image/jpeg;base64,{b64}' style='width:100%; max-height:95px; object-fit:cover; border-radius:5px; border:1px solid #cbd5e1;'/></div>"
            except Exception:
                pass
        
        prio_badge_bg = '#fef2f2' if priorite == 'urgent' else '#fffbeb' if priorite == 'moyen' else '#f0fdf4'
        prio_badge_col = '#dc2626' if priorite == 'urgent' else '#d97706' if priorite == 'moyen' else '#16a34a'
        
        # Aperçu interactif au survol (Tooltip avec photo)
        tooltip_html = f"""
        <div style="font-family: 'Inter', system-ui, sans-serif; width: 200px; padding: 2px;">
            {photo_mini}
            <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:3px;">
                <span style="font-size:13px; font-weight:800; color:#0f172a;">Dépôt #{depot_id}</span>
                <span style="background:{'#ecfdf5' if is_resolu else prio_badge_bg}; color:{'#047857' if is_resolu else prio_badge_col}; padding:1px 6px; border-radius:8px; font-size:9.5px; font-weight:700;">
                    {'✅ RÉSOLU' if is_resolu else priorite.upper()}
                </span>
            </div>
            <div style="font-size:11px; color:#2563eb; font-weight:700; margin-bottom:2px;">{commune_nom} • Qt. {quartier_nom}</div>
            <div style="font-size:11px; color:#334155;">Volume : <b>{volume:.2f} m³</b> {'(évacués)' if is_resolu else f'(~{bennes_estimees} camions)'}</div>
            <div style="font-size:9.5px; color:#94a3b8; margin-top:4px; border-top:1px dashed #e2e8f0; padding-top:3px; text-align:center;">
                👆 Cliquer pour ouvrir la fiche complète
            </div>
        </div>
        """
        
        if is_resolu:
            popup_text = f"""
            <div style="font-family: 'Inter', system-ui, -apple-system, sans-serif; min-width: 230px; max-width: 270px; padding: 2px;">
                <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:6px; border-bottom:1px solid #e2e8f0; padding-bottom:4px;">
                    <span style="font-size:14px; font-weight:800; color:#0f172a;">Dépôt #{depot_id}</span>
                    <span style="background:#ecfdf5; color:#047857; border:1px solid #a7f3d0; padding:2px 8px; border-radius:12px; font-size:10px; font-weight:700;">✅ RÉSOLU</span>
                </div>
                {photo_thumb}
                <div style="font-size:11px; color:#475569; margin-bottom:3px;">
                    <strong style="color:#0f172a;">🏛️ Commune :</strong> <span style="font-weight:700; color:#2563eb;">{commune_nom}</span>
                </div>
                <div style="font-size:11px; color:#475569; margin-bottom:4px;">
                    <strong style="color:#0f172a;">📍 Quartier exact :</strong> <span style="font-weight:700; color:#dc2626;">{quartier_nom}</span>
                </div>
                <div style="font-size:12px; color:#047857; margin-bottom:4px;">
                    <strong style="color:#1e293b;">Volume évacué :</strong> <span style="font-weight:700;">{volume:.2f} m³ traités</span>
                </div>
                <div style="font-size:11px; color:#64748b; margin-bottom:8px; line-height:1.3;">
                    <i class="fas fa-map-marker-alt" style="color:#94a3b8; margin-right:4px;"></i>{adresse}
                </div>
                <div style="margin-top:8px; padding-top:6px; border-top:1px solid #f1f5f9; display:flex; flex-direction:column; gap:5px;">
                    <div style="font-size:10px; color:#94a3b8; text-align:center; font-family:monospace; background:#f1f5f9; padding:3px 6px; border-radius:4px;">
                        📍 GPS : {lat:.5f}, {lon:.5f}
                    </div>
                    <a href="/liste" style="display:block; text-align:center; background:#047857; color:#ffffff; padding:6px 10px; border-radius:6px; font-size:11px; font-weight:600; text-decoration:none;">
                        📋 Voir dans l'Historique des Dépôts
                    </a>
                </div>
            </div>
            """
            
            folium.Marker(
                location=[lat, lon],
                popup=folium.Popup(popup_text, max_width=280),
                tooltip=folium.Tooltip(tooltip_html, sticky=True),
                icon=folium.Icon(color='green', icon='check', prefix='fa')
            ).add_to(couche_resolus)
            
            folium.Circle(
                location=[lat, lon],
                radius=max(30, min(100, int(volume * 10))),
                color='#10b981',
                fill=True,
                fill_opacity=0.10,
                weight=1.5,
                interactive=False
            ).add_to(couche_rayons)
            
        else:
            if not is_default_coords:
                heat_points.append([lat, lon, max(1.0, volume)])
                
            if is_default_coords:
                couleur = 'purple'
                badge_bg, badge_color = '#faf5ff', '#7c3aed'
                badge_loc = "<div style='margin-bottom:6px;'><span style='background:#faf5ff; color:#7c3aed; border:1px solid #e9d5ff; padding:2px 8px; border-radius:10px; font-size:10px; font-weight:700;'>📍 POSITION ESTIMÉE (CENTRE)</span></div>"
            else:
                couleur = couleurs.get(priorite, 'blue')
                badge_bg, badge_color = badge_colors.get(priorite, ('#eff6ff', '#2563eb'))
                badge_loc = ""
            
            popup_text = f"""
            <div style="font-family: 'Inter', system-ui, -apple-system, sans-serif; min-width: 230px; max-width: 270px; padding: 2px;">
                <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:6px; border-bottom:1px solid #e2e8f0; padding-bottom:4px;">
                    <span style="font-size:14px; font-weight:800; color:#0f172a;">Dépôt #{depot_id}</span>
                    <span style="background:{badge_bg}; color:{badge_color}; padding:2px 8px; border-radius:12px; font-size:10px; font-weight:700; text-transform:uppercase;">{priorite}</span>
                </div>
                {badge_loc}
                {photo_thumb}
                <div style="font-size:11px; color:#475569; margin-bottom:3px;">
                    <strong style="color:#0f172a;">🏛️ Commune :</strong> <span style="font-weight:700; color:#2563eb;">{commune_nom}</span>
                </div>
                <div style="font-size:11px; color:#475569; margin-bottom:4px;">
                    <strong style="color:#0f172a;">📍 Quartier exact :</strong> <span style="font-weight:700; color:#dc2626;">{quartier_nom}</span>
                </div>
                <div style="font-size:12px; color:#475569; margin-bottom:4px;">
                    <strong style="color:#1e293b;">Volume :</strong> <span style="font-weight:700; color:#0f172a;">{volume:.2f} m³</span>
                </div>
                <div style="font-size:12px; color:#475569; margin-bottom:6px;">
                    <strong style="color:#1e293b;">Statut :</strong> <span style="font-weight:600; color:#2563eb;">{statut_text}</span>
                </div>
                <div style="font-size:11px; color:#64748b; margin-bottom:8px; line-height:1.3;">
                    <i class="fas fa-map-marker-alt" style="color:#94a3b8; margin-right:4px;"></i>{adresse}
                </div>
                <div style="margin-top:8px; padding-top:6px; border-top:1px solid #f1f5f9; display:flex; flex-direction:column; gap:5px;">
                    <div style="display:flex; justify-content:space-between; align-items:center; background:#f8fafc; border:1px solid #e2e8f0; border-radius:6px; padding:4px 8px; font-size:11px;">
                        <span style="color:#64748b;">🚚 Besoin collecte :</span>
                        <span style="font-weight:700; color:#0f172a;">~{bennes_estimees} camion(s)</span>
                    </div>
                    <div style="font-size:10px; color:#94a3b8; text-align:center; font-family:monospace; background:#f1f5f9; padding:3px 6px; border-radius:4px;">
                        📍 GPS : {lat:.5f}, {lon:.5f}
                    </div>
                    <a href="/liste" style="display:block; text-align:center; background:#0f172a; color:#ffffff; padding:6px 10px; border-radius:6px; font-size:11px; font-weight:600; text-decoration:none;">
                        📋 Gérer dans la Liste des Dépôts
                    </a>
                </div>
            </div>
            """
            
            folium.Marker(
                location=[lat, lon],
                popup=folium.Popup(popup_text, max_width=280),
                tooltip=folium.Tooltip(tooltip_html, sticky=True),
                icon=folium.Icon(color=couleur, icon='trash', prefix='fa')
            ).add_to(couche_actifs)
            
            folium.Circle(
                location=[lat, lon],
                radius=max(35, min(150, int(volume * 14))),
                color=couleur,
                fill=True,
                fill_opacity=0.12,
                weight=1.5,
                interactive=False
            ).add_to(couche_rayons)
    
    if heat_points:
        heat_layer = folium.FeatureGroup(name="🔥 Carte thermique (Densité)", show=False)
        HeatMap(heat_points, radius=25, blur=15, min_opacity=0.3).add_to(heat_layer)
        heat_layer.add_to(carte)
        
    folium.LayerControl(position='topright', collapsed=True).add_to(carte)
    return carte._repr_html_()

# ==================== SIDEBAR ====================
sidebar = html.Div([
    html.Div([
        html.Div([
            html.Div([
                html.I(className="fas fa-leaf", style={'fontSize': '18px', 'color': '#10b981'})
            ], style={'width': '38px', 'height': '38px', 'borderRadius': '10px', 'backgroundColor': 'rgba(16, 185, 129, 0.15)', 'display': 'flex', 'alignItems': 'center', 'justifyContent': 'center', 'marginRight': '12px'}),
            html.Div([
                html.Span("SANUYA", style={'fontSize': '18px', 'fontWeight': '800', 'color': '#ffffff', 'letterSpacing': '0.5px'}),
                html.Span("PROPRETÉ URBAINE", style={'fontSize': '9px', 'fontWeight': '700', 'color': '#64748b', 'letterSpacing': '1.5px', 'display': 'block'})
            ])
        ], className="d-flex align-items-center px-3 py-3")
    ], style={'borderBottom': '1px solid rgba(255,255,255,0.08)'}),
    
    html.Div([
        html.P("NAVIGATION", className="text-white-50 px-3 mt-4 mb-2", style={'fontSize': '10px', 'fontWeight': '700', 'letterSpacing': '1.5px'}),
        dbc.Nav([
            dbc.NavLink([html.I(className="fas fa-chart-pie me-3"), html.Span("Tableau de bord")], href="/", active="exact", className="sidebar-link py-2 px-3 rounded-2 mb-1"),
            dbc.NavLink([html.I(className="fas fa-list-ul me-3"), html.Span("Liste des dépôts")], href="/liste", active="exact", className="sidebar-link py-2 px-3 rounded-2 mb-1"),
            dbc.NavLink([html.I(className="fas fa-chart-bar me-3"), html.Span("Statistiques & IA")], href="/stats", active="exact", className="sidebar-link py-2 px-3 rounded-2 mb-1"),
            dbc.NavLink([html.I(className="fas fa-camera-retro me-3"), html.Span("Tester l'IA")], href="/tester", active="exact", className="sidebar-link py-2 px-3 rounded-2 mb-1"),
        ], vertical=True, pills=False, className="px-2")
    ], className="flex-grow-1"),
    
    html.Div([
        html.Div([
            html.Div([
                html.Span(className="pulse-dot me-2"),
                html.Span("Système Actif (SQLite)", style={'fontSize': '11px', 'fontWeight': '600', 'color': '#cbd5e1'})
            ], className="d-flex align-items-center mb-1"),
            html.Div("Bamako, Mali • v2.1 Pro", style={'fontSize': '10px', 'color': '#64748b'})
        ], style={'padding': '12px 14px', 'background': 'rgba(255,255,255,0.03)', 'borderRadius': '8px', 'border': '1px solid rgba(255,255,255,0.06)'})
    ], className="p-3 mt-auto")
], style={
    'position': 'fixed', 'top': 0, 'left': 0, 'bottom': 0,
    'width': '235px', 'background': '#0f172a', 'padding': '0',
    'zIndex': '1000', 'display': 'flex', 'flexDirection': 'column'
})

# ==================== CONTENU ====================
content = html.Div([
    dcc.Location(id="url", refresh=False),
    html.Div(id="page-content"),
    dcc.Interval(id="interval-stats", interval=30000)
], style={'marginLeft': '235px', 'padding': '28px 36px', 'backgroundColor': '#f8fafc', 'minHeight': '100vh'})

app.layout = html.Div([sidebar, content])

# ==================== STAT CARD PRO ====================
def create_stat_card(title, value, unit="", icon="fas fa-trash", color="blue", subtitle=""):
    color_map = {
        'blue': {'bg': '#eff6ff', 'border': '#dbeafe', 'icon': '#2563eb'},
        'red': {'bg': '#fef2f2', 'border': '#fee2e2', 'icon': '#dc2626'},
        'orange': {'bg': '#fff7ed', 'border': '#ffedd5', 'icon': '#ea580c'},
        'green': {'bg': '#f0fdf4', 'border': '#dcfce7', 'icon': '#16a34a'},
    }
    c = color_map.get(color, color_map['blue'])
    return html.Div([
        html.Div([
            html.Div([
                html.Span(title, style={'fontSize': '12px', 'fontWeight': '700', 'color': '#64748b', 'textTransform': 'uppercase', 'letterSpacing': '0.5px'}),
                html.Div([
                    html.Span(str(value), style={'fontSize': '28px', 'fontWeight': '800', 'color': '#0f172a', 'lineHeight': '1.2'}),
                    html.Span(f" {unit}" if unit else "", style={'fontSize': '14px', 'fontWeight': '600', 'color': '#64748b', 'marginLeft': '4px'})
                ], className="d-flex align-items-baseline mt-1"),
                html.Div(subtitle, style={'fontSize': '11.5px', 'color': '#94a3b8', 'marginTop': '4px'}) if subtitle else html.Div()
            ]),
            html.Div([
                html.I(className=icon, style={'fontSize': '18px', 'color': c['icon']})
            ], style={'width': '42px', 'height': '42px', 'borderRadius': '10px', 'backgroundColor': c['bg'], 'border': f"1px solid {c['border']}", 'display': 'flex', 'alignItems': 'center', 'justifyContent': 'center'})
        ], className="d-flex justify-content-between align-items-start")
    ], className="stat-card col-md-3 col-sm-6")

# ==================== PAGE TABLEAU DE BORD ====================
def page_dashboard():
    return html.Div([
        html.Div([
            html.Div([
                html.H1("Tableau de bord stratégique", className="page-title"),
                html.P("Supervision et cartographie opérationnelle des dépôts sauvages en temps réel", className="page-subtitle"),
            ]),
            html.Div([
                dbc.Button([
                    html.I(className="fas fa-sync-alt me-2"),
                    "Actualiser la carte"
                ], id="btn-refresh-map", color="light", size="sm", className="border shadow-sm px-3 py-2 fw-semibold me-2", style={'fontSize': '12.5px'}),
                html.Span([
                    html.Span(className="pulse-dot me-2"),
                    "Données en direct"
                ], className="badge bg-success-subtle text-success border border-success-subtle px-3 py-2 d-inline-flex align-items-center", style={'fontSize': '12px', 'fontWeight': '600'})
            ], className="d-flex align-items-center mt-2 mt-md-0")
        ], className="d-flex flex-column flex-md-row justify-content-between align-items-start align-items-md-center mb-4"),
        
        html.Div(id="stats-container"),
        
        html.Div([
            html.Div([
                html.Div([
                    html.Div([
                        html.I(className="fas fa-map-marked-alt text-primary me-2", style={'fontSize': '18px'}),
                        html.H5("Système d'Information Géographique (SIG)", className="mb-0", style={'fontWeight': '700', 'color': '#0f172a'}),
                    ], className="d-flex align-items-center"),
                    html.Div([
                        html.Span("Multi-couches (OSM, Satellite, Relief)", className="badge bg-light text-muted border me-2", style={'fontSize': '11px'}),
                        html.Span("Haversine 50m actif", className="badge bg-primary-subtle text-primary border border-primary-subtle", style={'fontSize': '11px'})
                    ], className="d-flex align-items-center")
                ], className="d-flex justify-content-between align-items-center mb-3 pb-2 border-bottom"),
                
                dcc.Loading(
                    id="loading-map",
                    type="dot",
                    color="#0284c7",
                    children=html.Div(id="dashboard-map-container", className="map-container")
                ),
                
                html.Div([
                    html.Div([
                        html.Div([html.Span(className="legend-dot", style={'backgroundColor': '#dc2626'}), html.Span("Urgent (> 5 m³)", className="fw-semibold text-secondary")], className="legend-item"),
                        html.Div([html.Span(className="legend-dot", style={'backgroundColor': '#d97706'}), html.Span("Moyen (2 - 5 m³)", className="fw-semibold text-secondary")], className="legend-item"),
                        html.Div([html.Span(className="legend-dot", style={'backgroundColor': '#16a34a'}), html.Span("Normal (< 2 m³)", className="fw-semibold text-secondary")], className="legend-item"),
                        html.Div([html.Span(className="legend-dot", style={'backgroundColor': '#7c3aed'}), html.Span("Position estimée (Bamako centre)", className="fw-semibold text-secondary")], className="legend-item"),
                    ], className="d-flex flex-wrap gap-3 align-items-center"),
                    html.Small("💡 Astuce : Utilisez le sélecteur de couches en haut à droite pour basculer en vue Satellite ou Thermique.", className="text-muted")
                ], className="legend d-flex flex-column flex-md-row justify-content-between align-items-start align-items-md-center mt-3")
            ], className="content-card mt-3")
        ])
    ])

# ==================== PAGE LISTE ====================
def page_liste():
    return html.Div([
        html.H1("Liste des dépôts", className="page-title"),
        html.P("Gérez tous les signalements", className="page-subtitle mb-4"),
        html.Div([
            html.Div([
                html.Div([
                    html.Label("Filtrer par priorité"),
                    dcc.Dropdown(
                        id="filtre-priorite",
                        options=[
                            {'label': 'Tous', 'value': 'tous'},
                            {'label': 'Urgent', 'value': 'urgent'},
                            {'label': 'Moyen', 'value': 'moyen'},
                            {'label': 'Normal', 'value': 'normal'}
                        ],
                        value='tous',
                        className="dash-dropdown"
                    )
                ], className="filter-group"),
                html.Div([
                    html.Label("Filtrer par statut"),
                    dcc.Dropdown(
                        id="filtre-statut",
                        options=[
                            {'label': 'Tous les statuts (Actifs & Résolus)', 'value': 'tous'},
                            {'label': 'Dépôts actifs uniquement (À traiter)', 'value': 'actifs'},
                            {'label': 'En attente', 'value': 'en_attente'},
                            {'label': 'En cours de traitement', 'value': 'en_cours'},
                            {'label': 'Résolus / Traités', 'value': 'resolu'}
                        ],
                        value='tous',
                        className="dash-dropdown"
                    )
                ], className="filter-group"),
                html.Div([
                    html.Label("Filtrer par commune"),
                    dcc.Dropdown(
                        id="filtre-commune",
                        options=[
                            {'label': 'Toutes les communes', 'value': 'tous'},
                            {'label': 'Commune I (Djelibougou / Sotuba)', 'value': 'Commune I'},
                            {'label': 'Commune II (Médina-Coura / Hippodrome)', 'value': 'Commune II'},
                            {'label': 'Commune III (Centre / Badialan)', 'value': 'Commune III'},
                            {'label': 'Commune IV (Lafiabougou / Hamdallaye)', 'value': 'Commune IV'},
                            {'label': 'Commune V (Badalabougou / Baco-Djicoroni)', 'value': 'Commune V'},
                            {'label': 'Commune VI (Faladié / Sogoniko)', 'value': 'Commune VI'}
                        ],
                        value='tous',
                        className="dash-dropdown"
                    )
                ], className="filter-group"),
                html.Div([
                    html.Label("Rechercher par ID"),
                    dcc.Input(
                        id="search-id",
                        type="number",
                        placeholder="ID du dépôt",
                        style={'padding': '8px 12px', 'borderRadius': '8px', 'border': '1px solid #e9edf2', 'width': '150px'}
                    )
                ], className="filter-group"),
                html.Div([
                    html.Label("Actions & Exports"),
                    html.Div([
                        dbc.Button("Appliquer", id="btn-apply-filters", color="primary", size="sm", className="me-1"),
                        dbc.Button("Reset", id="btn-reset-filters", color="secondary", size="sm", className="me-2"),
                        html.A(
                            [html.I(className="fas fa-file-excel me-1"), "Excel (.xlsx)"],
                            id="btn-export-excel-link",
                            href="/export/excel",
                            download="sanuya_depots.xlsx",
                            target="_blank",
                            className="btn btn-sm btn-success shadow-sm me-1 text-white text-decoration-none d-inline-flex align-items-center"
                        ),
                        html.A(
                            [html.I(className="fas fa-file-pdf me-1"), "Rapport PDF"],
                            id="btn-export-pdf-link",
                            href="/export/pdf",
                            download="sanuya_rapport_intervention.pdf",
                            target="_blank",
                            className="btn btn-sm btn-danger shadow-sm me-1 text-white text-decoration-none d-inline-flex align-items-center"
                        ),
                        html.A(
                            [html.I(className="fas fa-file-csv me-1"), "CSV"],
                            id="btn-export-csv-link",
                            href="/export/csv",
                            download="sanuya_depots.csv",
                            target="_blank",
                            className="btn btn-sm btn-outline-secondary shadow-sm text-decoration-none d-inline-flex align-items-center"
                        ),
                    ], style={'display': 'flex', 'gap': '6px', 'flexWrap': 'wrap', 'alignItems': 'center'})
                ], className="filter-group"),
            ], className="filter-section"),
            html.Div(id="liste-cartes"),
            # Modale Statut
            dbc.Modal([
                dbc.ModalHeader("Changer le statut"),
                dbc.ModalBody([
                    html.P("ID du dépôt", style={'fontWeight': '500', 'marginBottom': '4px'}),
                    html.Div(id="modal-status-id", style={'fontSize': '18px', 'fontWeight': '700', 'marginBottom': '16px'}),
                    html.P("Nouveau statut", style={'fontWeight': '500', 'marginBottom': '4px'}),
                    dcc.Dropdown(
                        id="modal-status-select",
                        options=[
                            {'label': 'En attente', 'value': 'en_attente'},
                            {'label': 'En cours', 'value': 'en_cours'},
                            {'label': 'Résolu', 'value': 'resolu'}
                        ],
                        value='en_attente'
                    )
                ]),
                dbc.ModalFooter([
                    dbc.Button("Annuler", id="modal-status-close", className="ms-auto", n_clicks=0),
                    dbc.Button("Confirmer", id="modal-status-confirm", color="primary")
                ])
            ], id="modal-status", is_open=False),
            # Modale Priorité
            dbc.Modal([
                dbc.ModalHeader("Changer la priorité"),
                dbc.ModalBody([
                    html.P("ID du dépôt", style={'fontWeight': '500', 'marginBottom': '4px'}),
                    html.Div(id="modal-priorite-id", style={'fontSize': '18px', 'fontWeight': '700', 'marginBottom': '16px'}),
                    html.P("Nouvelle priorité", style={'fontWeight': '500', 'marginBottom': '4px'}),
                    dcc.Dropdown(
                        id="modal-priorite-select",
                        options=[
                            {'label': 'Urgent', 'value': 'urgent'},
                            {'label': 'Moyen', 'value': 'moyen'},
                            {'label': 'Normal', 'value': 'normal'}
                        ],
                        value='normal'
                    )
                ]),
                dbc.ModalFooter([
                    dbc.Button("Annuler", id="modal-priorite-close", className="ms-auto", n_clicks=0),
                    dbc.Button("Confirmer", id="modal-priorite-confirm", color="primary")
                ])
            ], id="modal-priorite", is_open=False),
            # Modale Suppression
            dbc.Modal([
                dbc.ModalHeader("Confirmer la suppression"),
                dbc.ModalBody([
                    html.P("Êtes-vous sûr de vouloir supprimer ce dépôt ?", style={'fontWeight': '500'}),
                    html.Div(id="modal-delete-id", style={'fontSize': '18px', 'fontWeight': '700', 'marginBottom': '16px'}),
                    html.P("Cette action est irréversible.", style={'color': '#ef4444', 'fontWeight': '500'})
                ]),
                dbc.ModalFooter([
                    dbc.Button("Annuler", id="modal-delete-close", className="ms-auto", n_clicks=0),
                    dbc.Button("Supprimer", id="modal-delete-confirm", color="danger")
                ])
            ], id="modal-delete", is_open=False),
            # Modale Photo
            dbc.Modal([
                dbc.ModalHeader("Photo du dépôt"),
                dbc.ModalBody([
                    html.Div(id="modal-photo-container", style={'textAlign': 'center'})
                ]),
                dbc.ModalFooter([
                    dbc.Button("Fermer", id="modal-photo-close", className="ms-auto", n_clicks=0)
                ])
            ], id="modal-photo", is_open=False, size="lg"),
            # Modale Carte
            dbc.Modal([
                dbc.ModalHeader("Localisation du dépôt"),
                dbc.ModalBody([
                    html.Div(id="modal-map-container", style={'width': '100%'})
                ]),
                dbc.ModalFooter([
                    dbc.Button("Fermer", id="modal-map-close", className="ms-auto", n_clicks=0)
                ])
            ], id="modal-map", is_open=False, size="xl")
        ], className="content-card")
    ])

# ==================== PAGE STATISTIQUES ====================
def page_stats():
    return html.Div([
        html.H1("Statistiques", className="page-title"),
        html.P("Analyse complète des données", className="page-subtitle mb-4"),
        html.Div([
            dbc.Row([
                dbc.Col([
                    dbc.Card([
                        dbc.CardHeader(html.H5("Répartition par priorité", className="mb-0")),
                        dbc.CardBody([
                            dcc.Graph(id="stats-priorite", config={'displayModeBar': False})
                        ])
                    ], className="shadow-sm border-0")
                ], md=6),
                dbc.Col([
                    dbc.Card([
                        dbc.CardHeader(html.H5("Répartition par statut", className="mb-0")),
                        dbc.CardBody([
                            dcc.Graph(id="stats-statut", config={'displayModeBar': False})
                        ])
                    ], className="shadow-sm border-0")
                ], md=6),
            ], className="mb-4"),
            dbc.Card([
                dbc.CardHeader(html.H5("Top 10 des dépôts par volume", className="mb-0")),
                dbc.CardBody([
                    dcc.Graph(id="stats-volume", config={'displayModeBar': False})
                ])
            ], className="shadow-sm border-0"),
            dcc.Interval(id="interval-stats-graphs", interval=30000)
        ])
    ])

# ==================== PAGE TESTER ====================
def page_tester():
    return html.Div([
        dcc.Store(id="store-batch-analyses", data=[]),
        html.Div([
            html.H1("Tester & Ingestion par Lot", className="page-title"),
            html.P("Analysez, inspectez et validez une ou plusieurs photos avant enregistrement dans Sanuya", className="page-subtitle mb-4"),
        ]),
        dbc.Row([
            dbc.Col([
                dbc.Card([
                    dbc.CardHeader(html.H5("1. Téléversement des photos", className="mb-0 fw-bold text-dark")),
                    dbc.CardBody([
                        dcc.Upload(
                            id="upload-photo",
                            children=html.Div([
                                html.I(className="fas fa-cloud-upload-alt", style={'fontSize': '42px', 'color': '#0284c7'}),
                                html.P("Glissez-déposez ou cliquez pour sélectionner des photos", className="mt-2 fw-semibold", style={'color': '#1e293b'}),
                                html.Span("Support multi-sélection (Batch) • JPEG, PNG", className="text-muted small")
                            ], className="upload-area p-3 text-center border-dashed rounded-3"),
                            multiple=True
                        ),
                        html.Div(id="upload-status", className="mt-2 text-center"),
                        html.Hr(className="my-3"),
                        html.H6("Coordonnées GPS de secours", className="mb-1 fw-bold text-dark"),
                        html.P("Appliquées seulement si la photo ne contient pas de géolocalisation EXIF automatique.", className="text-muted small mb-2"),
                        dbc.Row([
                            dbc.Col([
                                html.Label("Latitude", style={'fontSize': '12px', 'color': '#64748b'}),
                                dcc.Input(
                                    id="analyse-lat",
                                    type="number",
                                    value=12.6392,
                                    step=0.000001,
                                    className="form-control form-control-sm",
                                    style={'borderRadius': '6px'}
                                )
                            ], md=6),
                            dbc.Col([
                                html.Label("Longitude", style={'fontSize': '12px', 'color': '#64748b'}),
                                dcc.Input(
                                    id="analyse-lon",
                                    type="number",
                                    value=-8.0029,
                                    step=0.000001,
                                    className="form-control form-control-sm",
                                    style={'borderRadius': '6px'}
                                )
                            ], md=6),
                        ], className="g-2"),
                        dbc.Button(
                            html.Div([
                                html.I(className="fas fa-microchip me-2"),
                                "Lancer l'analyse & Vérification"
                            ]),
                            id="btn-analyser",
                            color="primary",
                            className="mt-3 w-100 fw-bold py-2 shadow-sm",
                            style={'borderRadius': '8px'}
                        )
                    ])
                ], className="shadow-sm border-0 rounded-3 mb-4")
            ], md=4),
            dbc.Col([
                dbc.Card([
                    dbc.CardHeader(html.H5("2. Console d'Inspection & Validation", className="mb-0 fw-bold text-dark")),
                    dbc.CardBody([
                        dcc.Loading(
                            id="loading-analyse",
                            type="dot",
                            color="#0284c7",
                            children=html.Div(id="analyse-result", className="text-center")
                        )
                    ])
                ], className="shadow-sm border-0 rounded-3")
            ], md=8),
        ])
    ])

# ==================== CALLBACKS DASHBOARD ====================
@app.callback(
    Output("stats-container", "children"),
    [Input("interval-stats", "n_intervals"),
     Input("url", "pathname")]
)
def update_stats_dashboard(n, pathname):
    if pathname not in ["/", ""]:
        return dash.no_update
    stats = get_stats_dashboard()
    if not stats:
        return html.Div("Impossible de charger les données", className="text-danger")
    
    total = stats.get('total', 0)
    resolu = stats.get('resolu', 0)
    total_historique = total + resolu
    taux_res = stats.get('taux_resolution', 0)
    
    stats_cards = html.Div([
        html.Div([
            create_stat_card("Total Dépôts", total, unit="actifs", icon="fas fa-layer-group", color="blue", subtitle=f"Sur {total_historique} signalements répertoriés"),
            create_stat_card("Priorité Urgente", stats.get('urgent', 0), unit="critiques", icon="fas fa-exclamation-triangle", color="red", subtitle="Nécessite évacuation immédiate"),
            create_stat_card("Volume Estimé", stats.get('volume_total', 0), unit="m³", icon="fas fa-cubes", color="orange", subtitle="Masse totale de déchets actifs"),
            create_stat_card("Taux Résolution", f"{taux_res}%", unit="", icon="fas fa-check-circle", color="green", subtitle=f"{resolu} résolus • {stats.get('cours', 0)} en cours"),
        ], className="row g-3 mb-3"),
        html.Div([
            html.Div([
                html.Div([
                    html.I(className="fas fa-clock text-warning me-2"),
                    html.Span("En attente : ", className="text-muted small"),
                    html.Strong(f"{stats.get('attente', 0)}", className="me-4 text-slate-800"),
                    html.I(className="fas fa-spinner text-primary me-2"),
                    html.Span("En cours : ", className="text-muted small"),
                    html.Strong(f"{stats.get('cours', 0)}", className="me-4 text-slate-800"),
                    html.I(className="fas fa-check-double text-success me-2"),
                    html.Span("Résolus : ", className="text-muted small"),
                    html.Strong(f"{resolu}", className="me-4 text-slate-800"),
                ], className="d-flex align-items-center flex-wrap"),
                html.Div([
                    html.I(className="fas fa-calendar-day me-2 text-muted"),
                    html.Span(f"Dernière synchronisation : {datetime.now().strftime('%d/%m/%Y • %H:%M')}", className="text-muted small")
                ], className="d-flex align-items-center mt-2 mt-md-0")
            ], className="d-flex flex-column flex-md-row justify-content-between align-items-start align-items-md-center py-2 px-3 bg-white rounded-3 border")
        ], className="mb-2")
    ])
    
    return stats_cards

@app.callback(
    Output("dashboard-map-container", "children"),
    [Input("url", "pathname"),
     Input("btn-refresh-map", "n_clicks")]
)
def update_dashboard_map(pathname, n_refresh):
    if pathname not in ["/", ""]:
        return dash.no_update
    map_html = generate_map()
    return html.Iframe(
        srcDoc=map_html,
        style={'width': '100%', 'height': '100%', 'border': 'none', 'borderRadius': '10px'}
    )

# ==================== CALLBACK LISTE ====================
@app.callback(
    Output("liste-cartes", "children"),
    [Input("btn-apply-filters", "n_clicks"),
     Input("btn-reset-filters", "n_clicks"),
     Input("modal-status-confirm", "n_clicks"),
     Input("modal-priorite-confirm", "n_clicks")],
    [State("filtre-priorite", "value"),
     State("filtre-statut", "value"),
     State("filtre-commune", "value"),
     State("search-id", "value")]
)
def update_liste_cartes(n_apply, n_reset, n_status_confirm, n_prio_confirm, priorite, statut, commune, search_id):
    ctx = dash.callback_context
    
    if not ctx.triggered:
        priorite = 'tous'
        statut = 'tous'
        commune = 'tous'
        search_id = None
    
    if ctx.triggered and ctx.triggered[0]['prop_id'].startswith('btn-reset-filters'):
        priorite = 'tous'
        statut = 'tous'
        commune = 'tous'
        search_id = None
    
    try:
        depots = get_depots_filtres(priorite, statut, commune)
    except Exception as e:
        print(f"❌ Erreur get_depots_filtres : {e}")
        depots = []
    
    if search_id:
        depots = [d for d in depots if d['id'] == search_id]
    
    if not depots:
        return html.Div("Aucun dépôt trouvé pour cette sélection", className="text-muted", style={'padding': '40px', 'textAlign': 'center'})
    
    total_vol = sum(d.get('volume', 0) or 0 for d in depots)
    total_bennes = max(1, int(round(total_vol / 5.0 + 0.49))) if total_vol > 0 else 0
    
    commune_info = f" • {commune}" if commune and commune != 'tous' else ""
    
    summary_bar = html.Div([
        html.Div([
            html.Span(f"{len(depots)} dépôt(s) affiché(s){commune_info}", className="fw-bold text-dark me-3"),
            html.Span("•", className="text-muted me-3"),
            html.Span([html.I(className="fas fa-layer-group text-primary me-1"), f"Volume cumulé : {total_vol:.2f} m³"], className="text-muted me-3"),
            html.Span("•", className="text-muted me-3"),
            html.Span([html.I(className="fas fa-truck-loading text-success me-1"), f"Logistique estimée : ~{total_bennes} benne(s) de 5 m³"], className="text-success fw-bold"),
        ], className="d-flex align-items-center flex-wrap small")
    ], className="p-2 px-3 mb-3 bg-white border rounded-3 shadow-sm")

    statut_labels = {'en_attente': 'En attente', 'en_cours': 'En cours', 'resolu': 'Résolu'}
    priorite_labels = {'urgent': 'Urgent', 'moyen': 'Moyen', 'normal': 'Normal'}

    cartes = []
    for d in depots:
        priorite_class = f"badge-{d['priorite']}" if d['priorite'] in ['urgent', 'moyen', 'normal'] else 'badge-normal'
        statut_class = f"badge-{d['statut']}" if d['statut'] in ['en_attente', 'en_cours', 'resolu'] else 'badge-en_attente'
        
        statut_text = statut_labels.get(d['statut'], str(d['statut']).capitalize())
        priorite_text = priorite_labels.get(d['priorite'], str(d['priorite']).capitalize())

        loc_info = get_location_details(d.get('latitude'), d.get('longitude'))
        commune_badge = d.get('commune') or loc_info['commune']
        quartier_badge = loc_info['quartier']
        repere_display = loc_info['repere'] if loc_info['repere'] else loc_info['adresse_complete']
        
        photo_chemin = d.get('photo_chemin', '')
        photo_existe = os.path.exists(photo_chemin) if photo_chemin else False
        
        vol = d.get('volume', 0) or 0
        bennes_depot = max(1, int(round(vol / 5.0 + 0.49))) if vol > 0 else 0

        cartes.append(html.Div([
            html.Div([
                html.Div([
                    html.Span(f"#{d['id']}", className="depot-id"),
                    html.Span(d['date'], className="depot-date"),
                    html.Span(f"{d['volume']} m³", className="depot-volume"),
                    html.Span([html.I(className="fas fa-truck me-1 text-muted"), f"~{bennes_depot} benne(s)"], className="badge bg-light text-dark border", style={'fontSize': '11px', 'fontWeight': '500'}),
                    html.Span([html.I(className="fas fa-landmark me-1 text-primary"), commune_badge], className="badge bg-primary-subtle text-primary border border-primary-subtle", style={'fontSize': '11px', 'fontWeight': '600'}),
                    html.Span([html.I(className="fas fa-map-pin me-1 text-danger"), f"Qt. {quartier_badge}"], className="badge bg-danger-subtle text-danger border border-danger-subtle", style={'fontSize': '11px', 'fontWeight': '700'}),
                    html.Span(priorite_text, className=f"badge {priorite_class}"),
                    html.Span(statut_text, className=f"badge {statut_class}"),
                    html.Span(repere_display, className="text-muted ms-2", style={'fontSize': '12px', 'maxWidth': '220px', 'overflow': 'hidden', 'textOverflow': 'ellipsis', 'whiteSpace': 'nowrap'}),
                ], className="depot-info"),
                html.Div([
                    html.Button([html.I(className="fas fa-camera me-1"), "Photo"], className="action-btn btn-photo",
                               id={'type': 'btn-photo', 'index': d['id']}, n_clicks=0,
                               style={'marginRight': '4px'}),
                    html.Button(
                        [html.I(className="fas fa-map-marked-alt me-1"), "Carte"],
                        className="action-btn btn-maps",
                        id={'type': 'btn-maps', 'index': d['id']},
                        n_clicks=0,
                        style={'marginRight': '4px'}
                    ),
                    html.Button([html.I(className="fas fa-tasks me-1"), "Statut"], className="action-btn btn-status", 
                               id={'type': 'btn-status', 'index': d['id']}, n_clicks=0,
                               style={'marginRight': '4px'}),
                    html.Button([html.I(className="fas fa-flag me-1"), "Priorité"], className="action-btn btn-priorite",
                               id={'type': 'btn-priorite', 'index': d['id']}, n_clicks=0,
                               style={'marginRight': '4px'}),
                    html.Button([html.I(className="fas fa-trash-alt me-1"), "Supprimer"], className="action-btn btn-delete",
                               id={'type': 'btn-delete', 'index': d['id']}, n_clicks=0),
                ], className="depot-actions")
            ], className="depot-card")
        ]))
    
    return [summary_bar] + cartes

# ==================== CALLBACKS MODALES ====================
@app.callback(
    Output("modal-status", "is_open"),
    Output("modal-status-id", "children"),
    Input({'type': 'btn-status', 'index': ALL}, 'n_clicks'),
    Input("modal-status-close", "n_clicks"),
    prevent_initial_call=True
)
def open_status_modal(status_clicks, close_clicks):
    ctx = dash.callback_context
    if not ctx.triggered:
        return False, ""
    
    trigger = ctx.triggered[0]
    trigger_val = trigger.get('value')
    if not trigger_val:
        return False, ""
    
    trigger_id = trigger['prop_id'].split('.')[0]
    if trigger_id == "modal-status-close":
        return False, ""
    
    try:
        btn_data = json.loads(trigger_id)
        depot_id = btn_data.get('index')
        return True, f"#{depot_id}"
    except:
        return False, ""

@app.callback(
    Output("modal-status", "is_open", allow_duplicate=True),
    Input("modal-status-confirm", "n_clicks"),
    State("modal-status-id", "children"),
    State("modal-status-select", "value"),
    prevent_initial_call=True
)
def confirm_status(n, depot_id, statut):
    if n is None or not depot_id:
        return False
    depot_id = int(depot_id.replace('#', ''))
    update_status_db(depot_id, statut)
    return False

@app.callback(
    Output("modal-priorite", "is_open"),
    Output("modal-priorite-id", "children"),
    Input({'type': 'btn-priorite', 'index': ALL}, 'n_clicks'),
    Input("modal-priorite-close", "n_clicks"),
    prevent_initial_call=True
)
def open_priorite_modal(priorite_clicks, close_clicks):
    ctx = dash.callback_context
    if not ctx.triggered:
        return False, ""
    
    trigger = ctx.triggered[0]
    trigger_val = trigger.get('value')
    if not trigger_val:
        return False, ""
    
    trigger_id = trigger['prop_id'].split('.')[0]
    if trigger_id == "modal-priorite-close":
        return False, ""
    
    try:
        btn_data = json.loads(trigger_id)
        depot_id = btn_data.get('index')
        return True, f"#{depot_id}"
    except:
        return False, ""

@app.callback(
    Output("modal-priorite", "is_open", allow_duplicate=True),
    Input("modal-priorite-confirm", "n_clicks"),
    State("modal-priorite-id", "children"),
    State("modal-priorite-select", "value"),
    prevent_initial_call=True
)
def confirm_priorite(n, depot_id, priorite):
    if n is None or not depot_id:
        return False
    depot_id = int(depot_id.replace('#', ''))
    update_priorite_db(depot_id, priorite)
    return False

@app.callback(
    Output("modal-delete", "is_open"),
    Output("modal-delete-id", "children"),
    Input({'type': 'btn-delete', 'index': ALL}, 'n_clicks'),
    Input("modal-delete-close", "n_clicks"),
    prevent_initial_call=True
)
def open_delete_modal(delete_clicks, close_clicks):
    ctx = dash.callback_context
    if not ctx.triggered:
        return False, ""
    
    trigger = ctx.triggered[0]
    trigger_val = trigger.get('value')
    if not trigger_val:
        return False, ""
    
    trigger_id = trigger['prop_id'].split('.')[0]
    if trigger_id == "modal-delete-close":
        return False, ""
    
    try:
        btn_data = json.loads(trigger_id)
        depot_id = btn_data.get('index')
        return True, f"#{depot_id}"
    except:
        return False, ""

@app.callback(
    Output("modal-delete", "is_open", allow_duplicate=True),
    Input("modal-delete-confirm", "n_clicks"),
    State("modal-delete-id", "children"),
    prevent_initial_call=True
)
def confirm_delete(n, depot_id):
    if n is None or not depot_id:
        return False
    depot_id = int(depot_id.replace('#', ''))
    delete_depot_db(depot_id)
    return False

# ==================== CALLBACK PHOTO ====================
@app.callback(
    Output("modal-photo", "is_open"),
    Output("modal-photo-container", "children"),
    Input({'type': 'btn-photo', 'index': ALL}, 'n_clicks'),
    Input("modal-photo-close", "n_clicks"),
    prevent_initial_call=True
)
def open_photo_modal(photo_clicks, close_clicks):
    ctx = dash.callback_context
    if not ctx.triggered:
        return False, html.Div()
    
    trigger = ctx.triggered[0]
    trigger_val = trigger.get('value')
    if not trigger_val:
        return False, html.Div()
    
    trigger_id = trigger['prop_id'].split('.')[0]
    if trigger_id == "modal-photo-close":
        return False, html.Div()
    
    try:
        btn_data = json.loads(trigger_id)
        depot_id = btn_data.get('index')
        
        conn = get_connection()
        if conn:
            cursor = conn.cursor()
            cursor.execute("SELECT photo_chemin, photo_nom FROM signalements WHERE id = %s", (depot_id,))
            result = cursor.fetchone()
            conn.close()
            
            if result and result[0]:
                photo_path = result[0]
                photo_nom = result[1] or "Photo"
                
                if os.path.exists(photo_path):
                    with open(photo_path, "rb") as f:
                        encoded = base64.b64encode(f.read()).decode()
                    src = f"data:image/jpeg;base64,{encoded}"
                    
                    return True, html.Div([
                        html.P(f"{photo_nom}", style={'textAlign': 'center', 'fontWeight': '600', 'marginBottom': '10px'}),
                        html.Img(src=src, style={'width': '100%', 'borderRadius': '8px', 'maxHeight': '600px', 'objectFit': 'contain'})
                    ])
                else:
                    return True, html.Div([
                        html.P("Photo non trouvée sur le serveur", className="text-center text-muted")
                    ])
            else:
                return True, html.Div([
                    html.P("Aucune photo associée à ce dépôt", className="text-center text-muted")
                ])
        return False, html.Div()
    except Exception as e:
        print(f"❌ Erreur photo : {e}")
        return False, html.Div()

# ==================== CALLBACK CARTE ====================
@app.callback(
    Output("modal-map", "is_open"),
    Output("modal-map-container", "children"),
    Input({'type': 'btn-maps', 'index': ALL}, 'n_clicks'),
    Input("modal-map-close", "n_clicks"),
    prevent_initial_call=True
)
def open_maps_modal(maps_clicks, close_clicks):
    ctx = dash.callback_context

    if not ctx.triggered:
        return False, html.Div()

    trigger = ctx.triggered[0]
    trigger_val = trigger.get('value')
    if not trigger_val:
        return False, html.Div()

    trigger_id = trigger['prop_id'].split('.')[0]
    if trigger_id == "modal-map-close":
        return False, html.Div()

    try:
        btn_data = json.loads(trigger_id)
        depot_id = btn_data.get('index')

        conn = get_connection()
        if not conn:
            return False, html.Div()

        cursor = conn.cursor()
        cursor.execute("""
            SELECT latitude, longitude, volume, priorite, statut
            FROM signalements
            WHERE id = %s
        """, (depot_id,))

        result = cursor.fetchone()
        conn.close()

        if not result:
            return True, html.Div([
                html.P("Dépôt introuvable", className="text-center text-muted")
            ])

        lat = float(result[0])
        lon = float(result[1])
        volume = result[2]
        priorite = result[3]
        statut = result[4]

        is_approximate = abs(lat - 12.6392) < 0.0005 and abs(lon - (-8.0029)) < 0.0005

        loc_info = get_location_details(lat, lon)
        commune_nom = loc_info['commune']
        quartier_nom = loc_info['quartier']
        adresse = loc_info['adresse_complete']
        if is_approximate and ("non disponible" in adresse.lower() or not adresse.strip()):
            adresse = "Bamako, Mali (Centre - Position estimée)"

        zoom_lvl = 14 if is_approximate else 17
        carte_depot = folium.Map(
            location=[lat, lon],
            zoom_start=zoom_lvl,
            tiles='OpenStreetMap'
        )

        folium.TileLayer(
            tiles='https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}',
            attr='Esri WorldImagery',
            name='Vue Satellite HD'
        ).add_to(carte_depot)
        folium.LayerControl(position='topright').add_to(carte_depot)

        couleurs = {
            'urgent': 'red',
            'moyen': 'orange',
            'normal': 'green'
        }
        couleur = 'purple' if is_approximate else couleurs.get(priorite, 'blue')
        bennes_estimees = max(1, round(volume / 5.0))

        popup_text = f"""
        <div style="font-family: 'Inter', system-ui, sans-serif; min-width: 210px; padding: 2px;">
            <div style="font-weight:800; font-size:14px; color:#0f172a; margin-bottom:4px;">Dépôt #{depot_id}</div>
            <div style="font-size:11px; color:#2563eb; font-weight:700;">{commune_nom} — Qt. {quartier_nom}</div>
            <div style="font-size:12px; margin:4px 0; color:#334155;">Volume : <b>{volume:.2f} m³</b> (~{bennes_estimees} camions)</div>
            <div style="font-size:11px; color:#64748b;">{adresse}</div>
        </div>
        """

        folium.Marker(
            location=[lat, lon],
            popup=folium.Popup(popup_text, max_width=280),
            tooltip=f"Dépôt #{depot_id} ({commune_nom})",
            icon=folium.Icon(color=couleur, icon='trash', prefix='fa')
        ).add_to(carte_depot)

        folium.Circle(
            location=[lat, lon],
            radius=100,
            color=couleur,
            fill=True,
            fill_opacity=0.15
        ).add_to(carte_depot)

        map_html = carte_depot._repr_html_()

        approx_banner = dbc.Alert([
            html.I(className="fas fa-info-circle me-2"),
            "Position approximative (centre de Bamako) : cette photo ne contenait pas de métadonnées GPS EXIF précises."
        ], color="warning", className="py-2 text-center mb-3", style={'fontSize': '13px'}) if is_approximate else html.Div()

        prio_color = 'danger' if priorite == 'urgent' else 'warning' if priorite == 'moyen' else 'success'

        return True, html.Div([
            approx_banner,
            html.Div([
                html.H5(
                    f"Dépôt #{depot_id} — Quartier {quartier_nom}",
                    style={'fontWeight': '800', 'marginBottom': '4px', 'color': '#0f172a'}
                ),
                html.P(
                    f"🏛️ {commune_nom} • {adresse}",
                    className="text-muted",
                    style={'marginBottom': '6px', 'fontSize': '13px'}
                ),
                html.Div([
                    dbc.Badge(f"Volume : {volume:.2f} m³", color="light", text_color="dark", className="me-2 py-1 px-2 border"),
                    dbc.Badge(f"Priorité : {priorite.upper()}", color=prio_color, className="me-2 py-1 px-2"),
                    dbc.Badge(f"Statut : {statut.capitalize()}", color="info", className="me-2 py-1 px-2"),
                    dbc.Badge(f"~{bennes_estimees} camion(s)", color="dark", className="py-1 px-2"),
                ], className="mb-1")
            ], className="text-center"),
            html.Hr(className="my-3"),
            html.Div(
                html.Iframe(
                    srcDoc=map_html,
                    width="100%",
                    height="520",
                    style={'border': 'none', 'borderRadius': '10px'}
                ),
                style={
                    'width': '100%',
                    'height': '520px',
                    'overflow': 'hidden',
                    'borderRadius': '10px',
                    'border': '1px solid #e9edf2'
                }
            ),
            html.Div([
                dcc.Link(
                    dbc.Button([
                        html.I(className="fas fa-layer-group me-2"),
                        "Explorer sur le SIG Principal"
                    ], color="primary", className="fw-semibold px-4 py-2", style={'borderRadius': '8px'}),
                    href="/"
                ),
                html.Span(
                    [html.I(className="fas fa-crosshairs me-1 text-primary"), f"GPS : {lat:.6f}, {lon:.6f}"],
                    className="badge bg-light text-dark border px-3 py-2 fw-semibold",
                    style={'fontSize': '12px', 'borderRadius': '8px'}
                )
            ], className="d-flex justify-content-center align-items-center flex-wrap gap-2 mt-3")
        ])

    except Exception as e:
        print(f"❌ Erreur carte dépôt : {e}")
        return True, html.Div([
            html.P(
                f"Erreur lors de l'affichage de la carte : {str(e)}",
                className="text-danger text-center"
            )
        ])

# ==================== UTILITAIRES D'EXPORT (EXCEL, PDF & CSV) ====================
def build_export_dataframe(depots):
    data_rows = []
    for d in depots:
        vol = d.get('volume', 0.0) or 0.0
        bennes = max(1, int(round(vol / 5.0 + 0.49))) if vol > 0 else 0
        lat, lon = d.get('latitude'), d.get('longitude')
        loc_info = get_location_details(lat, lon) if lat and lon else {'quartier': '-', 'commune': d.get('commune', 'Bamako'), 'repere': ''}
        maps_link = f"https://www.google.com/maps?q={lat},{lon}" if lat and lon else ""
        data_rows.append({
            "ID": d.get('id'),
            "Commune": d.get('commune') or loc_info['commune'],
            "Quartier exact": loc_info['quartier'],
            "Rue / Repère": loc_info['repere'],
            "Date de signalement": d.get('date', '-'),
            "Volume estimé (m³)": round(vol, 2),
            "Bennes 5m³ requises": bennes,
            "Priorité": str(d.get('priorite', '')).capitalize(),
            "Statut": str(d.get('statut', '')).replace('_', ' ').capitalize(),
            "Latitude": lat,
            "Longitude": lon,
            "Lien Itinéraire Google Maps": maps_link
        })
    return pd.DataFrame(data_rows)

def generate_styled_excel(df):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Dépôts Sanuya')
        ws = writer.sheets['Dépôts Sanuya']
        
        # En-tête bleu ardoise foncé
        header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
        header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        border_side = Side(style='thin', color='E2E8F0')
        border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        
        ws.row_dimensions[1].height = 26
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            
        # Zébrage et bordures
        fill_even = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=len(df)+1, min_col=1, max_col=len(df.columns)), start=2):
            ws.row_dimensions[row_idx].height = 20
            is_even = (row_idx % 2 == 0)
            for cell in row:
                cell.border = border
                cell.font = Font(name='Segoe UI', size=10)
                if is_even:
                    cell.fill = fill_even
                cell.alignment = Alignment(vertical='center')
                
        # Largeurs auto-ajustées
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    buf.seek(0)
    return buf.getvalue()

def generate_pdf_report(depots, filtre_commune='tous', filtre_statut='tous', filtre_priorite='tous'):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        rightMargin=20,
        leftMargin=20,
        topMargin=20,
        bottomMargin=20
    )
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=15, leading=18, textColor=colors.HexColor('#0f172a'), fontName='Helvetica-Bold')
    subtitle_style = ParagraphStyle('Sub', parent=styles['Normal'], fontSize=9, leading=12, textColor=colors.HexColor('#475569'))
    cell_style = ParagraphStyle('Cell', parent=styles['Normal'], fontSize=8, leading=10, textColor=colors.HexColor('#1e293b'))
    cell_bold = ParagraphStyle('CellBold', parent=styles['Normal'], fontSize=8, leading=10, textColor=colors.HexColor('#0f172a'), fontName='Helvetica-Bold')
    
    story = []
    story.append(Paragraph("RÉPUBLIQUE DU MALI • DISTRICT DE BAMAKO", subtitle_style))
    story.append(Paragraph("SANUYA - FICHE OFFICIELLE D'INTERVENTION & DE COLLECTE DES DÉPÔTS SAUVAGES", title_style))
    story.append(Spacer(1, 8))
    
    # Calculs KPI
    tot_vol = sum(d.get('volume', 0.0) or 0.0 for d in depots)
    tot_bennes = sum(max(1, int(round((d.get('volume', 0.0) or 0.0) / 5.0 + 0.49))) if (d.get('volume', 0.0) or 0.0) > 0 else 0 for d in depots)
    now_str = datetime.now().strftime('%d/%m/%Y à %H:%M')
    commune_txt = filtre_commune if filtre_commune != 'tous' else 'Toutes les Communes'
    statut_txt = str(filtre_statut).replace('_', ' ').capitalize() if filtre_statut != 'tous' else 'Tous (non résolus)'
    prio_txt = str(filtre_priorite).capitalize() if filtre_priorite != 'tous' else 'Toutes priorités'
    
    kpi_data = [
        [
            Paragraph(f"<b>Date d'émission :</b> {now_str}<br/><b>Périmètre :</b> {commune_txt}", subtitle_style),
            Paragraph(f"<b>Statut :</b> {statut_txt}<br/><b>Priorité :</b> {prio_txt}", subtitle_style),
            Paragraph(f"DÉPÔTS RECENSÉS<br/><font size='12' color='#0f172a'><b>{len(depots)}</b></font>", subtitle_style),
            Paragraph(f"VOLUME CUMULÉ<br/><font size='12' color='#2563eb'><b>{tot_vol:.1f} m³</b></font>", subtitle_style),
            Paragraph(f"BENNES 5m³ REQUISES<br/><font size='12' color='#16a34a'><b>~{tot_bennes} camions</b></font>", subtitle_style)
        ]
    ]
    kpi_table = Table(kpi_data, colWidths=[170, 160, 130, 140, 150])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#f1f5f9')),
        ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 10))
    
    table_data = [
        ['ID', 'Commune', 'Date Signalement', 'Priorité', 'Volume', 'Bennes', 'Statut', 'GPS', 'Itinéraire Maps']
    ]
    
    for d in depots:
        d_id = d.get('id')
        com = d.get('commune', 'Bamako')
        dt = d.get('date', '-')
        prio = str(d.get('priorite', '')).lower()
        vol = d.get('volume', 0.0) or 0.0
        bennes = max(1, int(round(vol / 5.0 + 0.49))) if vol > 0 else 0
        stat = str(d.get('statut', '')).replace('_', ' ').capitalize()
        lat, lon = d.get('latitude'), d.get('longitude')
        gps_txt = f"{lat:.4f}, {lon:.4f}" if lat and lon else '-'
        maps_url = f"https://maps.google.com/?q={lat},{lon}" if lat and lon else ''
        prio_color = '#dc2626' if prio == 'urgent' else '#d97706' if prio == 'moyen' else '#16a34a'
        
        table_data.append([
            Paragraph(f"<b>#{d_id}</b>", cell_bold),
            Paragraph(com, cell_style),
            Paragraph(str(dt), cell_style),
            Paragraph(f'<font color="{prio_color}"><b>{prio.upper()}</b></font>', cell_style),
            Paragraph(f"{vol:.1f} m³", cell_style),
            Paragraph(f"{bennes} benne(s)", cell_style),
            Paragraph(stat, cell_style),
            Paragraph(gps_txt, cell_style),
            Paragraph(f'<a href="{maps_url}" color="#2563eb"><u>Ouvrir Maps</u></a>' if maps_url else '-', cell_style)
        ])
        
    t = Table(table_data, colWidths=[40, 85, 95, 65, 65, 70, 75, 110, 145], repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0f172a')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 8),
        ('BOTTOMPADDING', (0,0), (-1,0), 5),
        ('TOPPADDING', (0,0), (-1,0), 5),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')]),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,1), (-1,-1), 4),
        ('BOTTOMPADDING', (0,1), (-1,-1), 4),
    ]))
    story.append(t)
    story.append(Spacer(1, 15))
    
    sig_data = [
        [
            Paragraph("<b>Le Chef d'Équipe Intervention & Voirie :</b>", subtitle_style),
            Paragraph("<b>La Direction des Services Urbains & Assainissement :</b>", subtitle_style)
        ],
        [
            Paragraph('<font color="#94a3b8">[Signature pour exécution]</font>', subtitle_style),
            Paragraph('<font color="#94a3b8">[Cachet officiel Mairie du District de Bamako]</font>', subtitle_style)
        ]
    ]
    sig_table = Table(sig_data, colWidths=[375, 375])
    sig_table.setStyle(TableStyle([
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
    ]))
    story.append(sig_table)
    
    doc.build(story)
    buf.seek(0)
    return buf.getvalue()

# ==================== ROUTES FLASK D'EXPORTATION DIRECTE ====================
# Les routes HTTP natives envoient l'en-tête Content-Disposition: attachment
# Cela garantit que Windows, Chrome et Edge conservent le nom officiel (ex: .xlsx, .pdf)
# sans jamais le renommer en GUID hexadécimal obscur.

@server.route('/export/excel')
def export_excel_route():
    try:
        priorite = request.args.get('priorite', 'tous')
        statut = request.args.get('statut', 'tous')
        commune = request.args.get('commune', 'tous')
        search_id = request.args.get('id', None)
        
        depots = get_depots_filtres(priorite, statut, commune)
        if search_id:
            try:
                sid_int = int(search_id)
                depots = [d for d in depots if d['id'] == sid_int]
            except Exception:
                pass
                
        df_export = build_export_dataframe(depots)
        excel_bytes = generate_styled_excel(df_export)
        date_str = datetime.now().strftime("%Y%m%d_%H%M")
        filename = f"sanuya_depots_{date_str}.xlsx"
        
        return Response(
            excel_bytes,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"; filename*=UTF-8\'\'{filename}',
                "Cache-Control": "no-cache, no-store, must-revalidate"
            }
        )
    except Exception as e:
        print(f"[ERREUR] Export Excel HTTP : {e}")
        return f"Erreur export Excel: {e}", 500

@server.route('/export/pdf')
def export_pdf_route():
    try:
        priorite = request.args.get('priorite', 'tous')
        statut = request.args.get('statut', 'tous')
        commune = request.args.get('commune', 'tous')
        search_id = request.args.get('id', None)
        
        depots = get_depots_filtres(priorite, statut, commune)
        if search_id:
            try:
                sid_int = int(search_id)
                depots = [d for d in depots if d['id'] == sid_int]
            except Exception:
                pass
                
        pdf_bytes = generate_pdf_report(
            depots,
            filtre_commune=commune,
            filtre_statut=statut,
            filtre_priorite=priorite
        )
        date_str = datetime.now().strftime("%Y%m%d_%H%M")
        filename = f"sanuya_rapport_intervention_{date_str}.pdf"
        
        return Response(
            pdf_bytes,
            mimetype="application/pdf",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"; filename*=UTF-8\'\'{filename}',
                "Cache-Control": "no-cache, no-store, must-revalidate"
            }
        )
    except Exception as e:
        print(f"[ERREUR] Export PDF HTTP : {e}")
        return f"Erreur export PDF: {e}", 500

@server.route('/export/csv')
def export_csv_route():
    try:
        priorite = request.args.get('priorite', 'tous')
        statut = request.args.get('statut', 'tous')
        commune = request.args.get('commune', 'tous')
        search_id = request.args.get('id', None)
        
        depots = get_depots_filtres(priorite, statut, commune)
        if search_id:
            try:
                sid_int = int(search_id)
                depots = [d for d in depots if d['id'] == sid_int]
            except Exception:
                pass
                
        df_export = build_export_dataframe(depots)
        csv_data = df_export.to_csv(index=False, sep=';', encoding='utf-8-sig')
        date_str = datetime.now().strftime("%Y%m%d_%H%M")
        filename = f"sanuya_depots_{date_str}.csv"
        
        return Response(
            csv_data,
            mimetype="text/csv; charset=utf-8-sig",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"; filename*=UTF-8\'\'{filename}',
                "Cache-Control": "no-cache, no-store, must-revalidate"
            }
        )
    except Exception as e:
        print(f"[ERREUR] Export CSV HTTP : {e}")
        return f"Erreur export CSV: {e}", 500

@server.route('/api/migrate')
def api_migrate_route():
    """Route web sécurisée permettant de déclencher ou vérifier la migration MySQL o2switch directement depuis le navigateur."""
    try:
        from database import init_mysql_db, get_connection
        succes, message = init_mysql_db()
        
        # Récupération du nombre de lignes en base
        count = 0
        try:
            conn = get_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM signalements")
            count = cursor.fetchone()[0]
            cursor.close()
            conn.close()
        except Exception:
            pass
            
        import json as py_json
        return Response(
            py_json.dumps({
                "succes": succes,
                "message": message,
                "total_signalements": count,
                "base_active": "MySQL o2switch (vuxe8870_sanuya)"
            }, ensure_ascii=False, indent=2),
            mimetype="application/json; charset=utf-8"
        )
    except Exception as e:
        import json as py_json
        return Response(
            py_json.dumps({"succes": False, "erreur": str(e)}),
            status=500,
            mimetype="application/json; charset=utf-8"
        )

# ==================== CALLBACK SYNCHRONISATION URLS EXPORT ====================

@app.callback(
    [Output("btn-export-excel-link", "href"),
     Output("btn-export-pdf-link", "href"),
     Output("btn-export-csv-link", "href")],
    [Input("filtre-priorite", "value"),
     Input("filtre-statut", "value"),
     Input("filtre-commune", "value"),
     Input("search-id", "value")]
)
def update_export_urls(priorite, statut, commune, search_id):
    query_parts = []
    if priorite and priorite != 'tous':
        query_parts.append(f"priorite={priorite}")
    if statut and statut != 'tous':
        query_parts.append(f"statut={statut}")
    if commune and commune != 'tous':
        query_parts.append(f"commune={commune}")
    if search_id:
        query_parts.append(f"id={search_id}")
    
    qs = ("?" + "&".join(query_parts)) if query_parts else ""
    return f"/export/excel{qs}", f"/export/pdf{qs}", f"/export/csv{qs}"

# ==================== CALLBACK UPLOAD ====================
@app.callback(
    Output("upload-status", "children"),
    Input("upload-photo", "contents"),
    State("upload-photo", "filename")
)
def handle_upload(contents, filename):
    if contents is None:
        return ""
    if isinstance(contents, list):
        return html.Span(f"📁 {len(contents)} photo(s) prête(s) pour l'analyse par lot", className="text-success fw-bold")
    return html.Span(f"📁 Fichier chargé : {filename}", style={'color': '#22c55e', 'fontWeight': '600'})

def traiter_une_photo_unitaire(content_string, orig_filename, lat_manuel, lon_manuel):
    decoded = base64.b64decode(content_string)
    os.makedirs("images_test", exist_ok=True)
    nom = f"analyse_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}.jpg"
    chemin = os.path.join("images_test", nom)
    with open(chemin, 'wb') as f:
        f.write(decoded)
    
    gps_data = get_gps_from_exif(chemin)
    gps_precision = get_gps_precision(chemin)
    location_metadata = get_location_from_metadata(chemin)
    
    location_coords = None
    if location_metadata:
        loc_name = location_metadata.replace('Lieu détecté: ', '').strip()
        if loc_name and len(loc_name) > 2:
            location_coords = search_coordinates_by_name(loc_name)
    
    if gps_data:
        lat = gps_data['latitude']
        lon = gps_data['longitude']
        gps_detecte = True
        precision = gps_precision if gps_precision else 5.0
        gps_source = "GPS EXIF certifié"
    elif location_coords:
        lat = location_coords['latitude']
        lon = location_coords['longitude']
        gps_detecte = False
        precision = None
        gps_source = "Recherche de lieu"
    else:
        lat = lat_manuel if lat_manuel is not None else 12.6392
        lon = lon_manuel if lon_manuel is not None else -8.0029
        gps_detecte = False
        precision = None
        gps_source = "Coordonnées par défaut"
    
    loc_info = get_location_details(lat, lon)
    commune_nom = loc_info['commune']
    quartier_nom = loc_info['quartier']
    adresse = loc_info['adresse_complete']
    
    resultat = analyser_photo(chemin)
    priorite = resultat.get('priorite', 'normal')
    vol = resultat.get('volume', 0.0)
    nb_dechets = resultat.get('nb', 0)
    bennes = max(1, int(round(vol / 5.0 + 0.49))) if vol > 0 else 0
    
    existants = get_depots_filtres('tous', 'tous', 'tous')
    is_dup, id_dup, distance = est_doublon(lat, lon, existants, seuil=50)
    
    return {
        'nom': nom,
        'chemin': chemin,
        'orig_filename': orig_filename,
        'lat': lat,
        'lon': lon,
        'commune': commune_nom,
        'quartier': quartier_nom,
        'adresse': adresse,
        'gps_detecte': gps_detecte,
        'gps_source': gps_source,
        'precision': precision,
        'resultat': resultat,
        'priorite': priorite,
        'volume': vol,
        'bennes': bennes,
        'nb_dechets': nb_dechets,
        'is_doublon': is_dup,
        'id_doublon': id_dup,
        'distance_doublon': distance,
        'image_b64': resultat.get('image_b64', base64.b64encode(decoded).decode('utf-8'))
    }

# ==================== CALLBACKS ANALYSE & VALIDATION PAR LOT ====================
@app.callback(
    [Output("analyse-result", "children"),
     Output("store-batch-analyses", "data")],
    Input("btn-analyser", "n_clicks"),
    [State("upload-photo", "contents"), 
     State("upload-photo", "filename"),
     State("analyse-lat", "value"), 
     State("analyse-lon", "value")],
    prevent_initial_call=True
)
def analyser_callback(n, contents, filename, lat_manuel, lon_manuel):
    if n is None or contents is None:
        return html.Div([
            html.I(className="fas fa-info-circle", style={'fontSize': '48px', 'color': '#94a3b8'}),
            html.P("Sélectionnez une ou plusieurs photos et cliquez sur Lancer l'analyse", className="text-muted mt-3", style={'fontSize': '15px'})
        ], className="py-5"), []
    
    try:
        if isinstance(contents, list):
            file_list = contents
            names_list = filename if isinstance(filename, list) else [filename]*len(contents)
        else:
            file_list = [contents]
            names_list = [filename]
        
        traites = []
        for c, fn in zip(file_list, names_list):
            if not c or ',' not in c:
                continue
            _, c_str = c.split(',')
            info = traiter_une_photo_unitaire(c_str, fn, lat_manuel, lon_manuel)
            traites.append(info)
        
        if not traites:
            return html.Div("Aucune image valide à traiter.", className="text-danger p-3"), []
        
        total_photos = len(traites)
        total_dechets = sum(t['nb_dechets'] for t in traites)
        total_vol = sum(t['volume'] for t in traites)
        total_bennes = max(1, int(round(total_vol / 5.0 + 0.49))) if total_vol > 0 else 0
        total_doublons = sum(1 for t in traites if t['is_doublon'])
        total_valides_prets = sum(1 for t in traites if t['nb_dechets'] > 0 and not t['is_doublon'])
        
        cartes_resultats = []
        
        # Barre de commande & validation supérieure
        batch_header = html.Div([
            html.Div([
                html.Div([
                    html.H5([
                        html.I(className="fas fa-tasks text-primary me-2"),
                        f"Console de Contrôle Qualité ({total_photos} photo(s) analysée(s))"
                    ], className="fw-bold mb-1 text-dark"),
                    html.P("Vérifiez les détections IA, le volume et la géolocalisation de chaque photo ci-dessous avant d'enregistrer.", className="text-muted small mb-2")
                ]),
                html.Div([
                    html.Span([html.I(className="fas fa-trash-alt me-1"), f"{total_dechets} déchet(s)"], className="badge bg-primary me-2 py-2 px-3"),
                    html.Span([html.I(className="fas fa-cubes me-1"), f"{total_vol:.2f} m³ cumulés"], className="badge bg-info text-dark me-2 py-2 px-3"),
                    html.Span([html.I(className="fas fa-truck me-1"), f"~{total_bennes} benne(s)"], className="badge bg-light text-dark border me-2 py-2 px-3"),
                    html.Span([html.I(className="fas fa-clone me-1"), f"{total_doublons} doublon(s)"], className="badge bg-warning text-dark me-2 py-2 px-3") if total_doublons > 0 else html.Span(),
                ], className="d-flex flex-wrap gap-1 mb-3")
            ]),
            html.Div([
                dbc.Button([
                    html.I(className="fas fa-save me-2"),
                    "Enregistrer les dépôts cochés dans Sanuya"
                ], id="btn-sauvegarder-selection", color="success", className="fw-bold px-4 py-2 shadow-sm me-2", style={'borderRadius': '8px'}),
                dbc.Button([
                    html.I(className="fas fa-check-square me-2"),
                    "Inverser la sélection"
                ], id="btn-toggle-all", color="light", size="sm", className="border fw-semibold px-3 py-2", style={'borderRadius': '8px'}),
            ], className="d-flex align-items-center flex-wrap gap-2"),
            html.Div(id="msg-sauvegarde-selection", className="w-100 mt-2")
        ], className="p-3 mb-4 bg-white rounded-3 border shadow-sm")
        cartes_resultats.append(batch_header)
        
        for i, t in enumerate(traites):
            prio = t['priorite']
            prio_color = '#ef4444' if prio == 'urgent' else '#f59e0b' if prio == 'moyen' else '#22c55e'
            default_check = (t['nb_dechets'] > 0 and not t['is_doublon'])
            
            status_badge = html.Span("✅ Prêt pour validation", className="badge bg-success py-1 px-2") if default_check else \
                           html.Span("⚠️ Doublon potentiel", className="badge bg-warning text-dark py-1 px-2") if t['is_doublon'] else \
                           html.Span("❌ Aucun déchet détecté", className="badge bg-secondary py-1 px-2")
            
            card = dbc.Card([
                dbc.CardHeader([
                    html.Div([
                        dbc.Checkbox(
                            id={'type': 'check-valid-depot', 'index': i},
                            value=default_check,
                            label=html.Span([
                                html.Strong(f"Photo #{i+1} : ", className="text-dark me-1"),
                                html.Span(t['orig_filename'], className="text-muted small")
                            ]),
                            className="fw-semibold fs-6 mb-0 d-flex align-items-center"
                        ),
                        status_badge
                    ], className="d-flex justify-content-between align-items-center w-100")
                ], className="bg-light py-2 px-3 border-bottom"),
                dbc.CardBody([
                    dbc.Row([
                        dbc.Col([
                            html.Div([
                                html.Img(
                                    src=f"data:image/jpeg;base64,{t['image_b64']}",
                                    style={'width': '100%', 'borderRadius': '8px', 'maxHeight': '260px', 'objectFit': 'contain', 'backgroundColor': '#0f172a'}
                                )
                            ], className="text-center mb-2")
                        ], md=5),
                        dbc.Col([
                            html.Div([
                                html.Div([
                                    html.Span([html.I(className="fas fa-landmark me-1 text-primary"), t['commune']], className="badge bg-primary-subtle text-primary border border-primary-subtle me-2 py-1 px-2"),
                                    html.Span([html.I(className="fas fa-map-pin me-1 text-danger"), f"Qt. {t['quartier']}"], className="badge bg-danger-subtle text-danger border border-danger-subtle me-2 py-1 px-2"),
                                    html.Span([html.I(className="fas fa-cubes me-1 text-info"), f"{t['volume']:.2f} m³ (~{t['bennes']} camions)"], className="badge bg-light text-dark border me-2 py-1 px-2"),
                                    html.Span(prio.upper(), className="badge py-1 px-2", style={'backgroundColor': prio_color}),
                                ], className="mb-2 d-flex flex-wrap gap-1"),
                                html.Div([
                                    html.I(className="fas fa-crosshairs me-1 text-primary"),
                                    html.Span(f"{t['gps_source']} : ", className="fw-bold small text-dark"),
                                    html.Span(f"{t['lat']:.5f}, {t['lon']:.5f}", className="font-monospace small text-muted"),
                                ], className="mb-2"),
                                html.Div([
                                    html.I(className="fas fa-map-marker-alt me-1 text-muted"),
                                    html.Span(t['adresse'], className="small text-muted")
                                ], className="mb-3"),
                            ]),
                            
                            # Ventilation des produits détectés
                            html.Div([
                                html.Span("Déchets & matières segmentés par l'IA :", className="small fw-bold text-dark d-block mb-1"),
                                html.Div([
                                    html.Span(f"{p_name} ×{p_qty}", className="badge me-1 mb-1 py-1 px-2", style={'backgroundColor': categoriser_produit(p_name)[1], 'fontSize': '11px'})
                                    for p_name, p_qty in t['resultat'].get('produits_counts', {}).items()
                                ] if t['resultat'].get('produits_counts') else [
                                    html.Span("Aucun type de matière identifié", className="text-muted small fst-italic")
                                ], className="d-flex flex-wrap gap-1 mb-2")
                            ], className="p-2 bg-light rounded-2 border mb-2"),
                            
                            dbc.Alert([
                                html.I(className="fas fa-exclamation-triangle me-2 text-warning"),
                                f"Doublon potentiel : dépôt situé à {t['distance_doublon']:.0f}m du Dépôt #{t['id_doublon']}. Décoché par sécurité."
                            ], color="warning", className="p-2 small mb-0") if t['is_doublon'] else html.Div(),
                            
                            dbc.Alert([
                                html.I(className="fas fa-info-circle me-2 text-secondary"),
                                "Aucun déchet visible n'a été confirmé sur ce cliché."
                            ], color="secondary", className="p-2 small mb-0") if t['nb_dechets'] == 0 else html.Div(),
                        ], md=7)
                    ])
                ], className="p-3")
            ], className="mb-3 shadow-sm border rounded-3")
            cartes_resultats.append(card)
        
        return html.Div(cartes_resultats), traites
    except Exception as e:
        return html.Div([
            html.I(className="fas fa-exclamation-triangle", style={'fontSize': '32px', 'color': '#ef4444'}),
            html.P(f"Erreur d'analyse : {str(e)}", className="text-danger mt-2 fw-semibold")
        ], className="text-center py-4"), []

# Callback de validation et enregistrement effectif en base
@app.callback(
    Output("msg-sauvegarde-selection", "children"),
    Input("btn-sauvegarder-selection", "n_clicks"),
    [State("store-batch-analyses", "data"),
     State({'type': 'check-valid-depot', 'index': ALL}, 'value'),
     State({'type': 'check-valid-depot', 'index': ALL}, 'id')],
    prevent_initial_call=True
)
def sauvegarder_lot_callback(n, analyses_data, check_values, check_ids):
    if not n or not analyses_data:
        return dash.no_update
        
    checked_indices = set()
    for val, cid in zip(check_values, check_ids):
        if val:
            checked_indices.add(cid['index'])
            
    if not checked_indices:
        return dbc.Alert([
            html.I(className="fas fa-exclamation-circle me-2"),
            "Aucun dépôt n'est coché. Veuillez cocher au moins un dépôt à enregistrer."
        ], color="warning", className="mt-2 fw-semibold")
        
    enregistres = 0
    for i, t in enumerate(analyses_data):
        if i in checked_indices:
            lat = t['lat']
            lon = t['lon']
            vol = t['volume']
            prio = t['priorite']
            nom = t['nom']
            chemin = t.get('chemin')
            res = add_detection(lat, lon, vol, prio, 'en_attente', nom, chemin)
            if res.get('success'):
                enregistres += 1
                
    return dbc.Alert([
        html.H5([
            html.I(className="fas fa-check-circle text-success me-2"),
            f"{enregistres} dépôt(s) validé(s) et enregistré(s) avec succès !"
        ], className="mb-2 fw-bold text-success"),
        html.P("Les dépôts vérifiés ont été ajoutés à la base de données et sont consultables immédiatement sur la carte SIG et dans la liste des dépôts.", className="mb-3 small text-dark"),
        html.Div([
            dcc.Link(dbc.Button([html.I(className="fas fa-layer-group me-2"), "Voir sur le SIG Principal"], color="primary", size="sm", className="me-2 fw-semibold px-3 py-2"), href="/"),
            dcc.Link(dbc.Button([html.I(className="fas fa-list me-2"), "Consulter la Liste des Dépôts"], color="light", size="sm", className="border fw-semibold px-3 py-2"), href="/liste")
        ])
    ], color="success", className="p-3 mt-3 shadow-sm rounded-3 border-success")

# Callback pour inverser la sélection
@app.callback(
    Output({'type': 'check-valid-depot', 'index': ALL}, 'value'),
    Input("btn-toggle-all", "n_clicks"),
    State({'type': 'check-valid-depot', 'index': ALL}, 'value'),
    prevent_initial_call=True
)
def toggle_all_checks(n, current_values):
    if not n or not current_values:
        return dash.no_update
    all_true = all(current_values)
    return [not all_true for _ in current_values]

# ==================== CALLBACK STATS GRAPHIQUES ====================
@app.callback(
    [Output("stats-priorite", "figure"),
     Output("stats-statut", "figure"),
     Output("stats-volume", "figure")],
    Input("interval-stats-graphs", "n_intervals")
)
def update_stats_graphs(n):
    def create_empty_figure(message="Aucune donnée disponible"):
        fig = go.Figure()
        fig.add_annotation(
            text=message,
            x=0.5,
            y=0.5,
            showarrow=False,
            font=dict(size=16, color="#64748b")
        )
        fig.update_layout(
            plot_bgcolor='rgba(0,0,0,0)',
            paper_bgcolor='rgba(0,0,0,0)',
            height=350,
            margin=dict(l=20, r=20, t=40, b=20)
        )
        return fig
    
    try:
        conn = get_connection()
        if not conn:
            fig_empty = create_empty_figure("Impossible de se connecter à la base de données")
            return fig_empty, fig_empty, fig_empty
        
        try:
            df = pd.read_sql("""
                SELECT 
                    priorite,
                    statut,
                    volume,
                    id
                FROM signalements 
                ORDER BY date_creation DESC
            """, conn)
            conn.close()
        except Exception as e:
            print(f"❌ Erreur lecture SQL: {e}")
            fig_empty = create_empty_figure(f"Erreur de lecture: {str(e)}")
            return fig_empty, fig_empty, fig_empty
        
        if df.empty:
            fig_empty = create_empty_figure("Aucune donnée dans la base")
            return fig_empty, fig_empty, fig_empty
        
        # --- Graphique 1: Priorité ---
        try:
            priorite_counts = df['priorite'].value_counts().reset_index()
            priorite_counts.columns = ['priorite', 'count']
            
            priorite_labels = {'urgent': 'Urgent', 'moyen': 'Moyen', 'normal': 'Normal'}
            priorite_counts['label'] = priorite_counts['priorite'].map(priorite_labels).fillna(priorite_counts['priorite'])
            
            couleurs_priorite = {'urgent': '#ef4444', 'moyen': '#f59e0b', 'normal': '#22c55e'}
            couleurs = [couleurs_priorite.get(p, '#94a3b8') for p in priorite_counts['priorite']]
            
            fig_priorite = go.Figure()
            if not priorite_counts.empty:
                fig_priorite.add_trace(go.Pie(
                    labels=priorite_counts['label'],
                    values=priorite_counts['count'],
                    marker=dict(colors=couleurs, line=dict(color='#ffffff', width=2)),
                    hole=0.55,
                    textposition='inside',
                    textinfo='percent+label',
                    hoverinfo='label+value+percent'
                ))
            else:
                fig_priorite.add_annotation(text="Aucune donnée", x=0.5, y=0.5, showarrow=False, font=dict(size=14, color="#94a3b8"))
            
            fig_priorite.update_layout(
                title=dict(text="Répartition par niveau d'urgence", font=dict(size=15, color="#0f172a", family="Inter")),
                plot_bgcolor='rgba(0,0,0,0)',
                paper_bgcolor='rgba(0,0,0,0)',
                font_color='#0f172a',
                height=340,
                margin=dict(l=20, r=20, t=50, b=20)
            )
        except Exception as e:
            print(f"❌ Erreur graphique priorité: {e}")
            fig_priorite = create_empty_figure("Erreur graphique priorité")
        
        # --- Graphique 2: Statut ---
        try:
            statut_counts = df['statut'].value_counts().reset_index()
            statut_counts.columns = ['statut', 'count']
            
            statut_labels = {'en_attente': 'En attente', 'en_cours': 'En cours', 'resolu': 'Résolu'}
            statut_counts['label'] = statut_counts['statut'].map(statut_labels).fillna(statut_counts['statut'])
            
            couleurs_statut = {'en_attente': '#f59e0b', 'en_cours': '#3b82f6', 'resolu': '#10b981'}
            statut_couleurs = [couleurs_statut.get(s, '#94a3b8') for s in statut_counts['statut']]
            
            fig_statut = go.Figure()
            if not statut_counts.empty:
                fig_statut.add_trace(go.Pie(
                    labels=statut_counts['label'],
                    values=statut_counts['count'],
                    marker=dict(colors=statut_couleurs, line=dict(color='#ffffff', width=2)),
                    hole=0.55,
                    textposition='inside',
                    textinfo='percent+label',
                    hoverinfo='label+value+percent'
                ))
            else:
                fig_statut.add_annotation(text="Aucune donnée", x=0.5, y=0.5, showarrow=False, font=dict(size=14, color="#94a3b8"))
            
            fig_statut.update_layout(
                title=dict(text="Avancement de l'éradication", font=dict(size=15, color="#0f172a", family="Inter")),
                plot_bgcolor='rgba(0,0,0,0)',
                paper_bgcolor='rgba(0,0,0,0)',
                font_color='#0f172a',
                height=340,
                margin=dict(l=20, r=20, t=50, b=20)
            )
        except Exception as e:
            print(f"❌ Erreur graphique statut: {e}")
            fig_statut = create_empty_figure("Erreur graphique statut")
        
        # --- Graphique 3: Volume ---
        try:
            top_volumes = df.nlargest(10, 'volume')[['id', 'volume']]
            
            fig_volume = go.Figure()
            if not top_volumes.empty:
                fig_volume.add_trace(go.Bar(
                    x=top_volumes['id'].astype(str),
                    y=top_volumes['volume'],
                    marker=dict(
                        color=top_volumes['volume'],
                        colorscale='Blues',
                        showscale=False
                    ),
                    text=top_volumes['volume'].round(2),
                    textposition='outside',
                    texttemplate='%{text:.2f} m³'
                ))
            else:
                fig_volume.add_annotation(text="Aucune donnée", x=0.5, y=0.5, showarrow=False, font=dict(size=14, color="#94a3b8"))
            
            fig_volume.update_layout(
                title=dict(text="Top 10 des dépôts les plus volumineux (m³)", font=dict(size=15, color="#0f172a", family="Inter")),
                plot_bgcolor='rgba(0,0,0,0)',
                paper_bgcolor='rgba(0,0,0,0)',
                font_color='#0f172a',
                xaxis_title="Identifiant du dépôt (#ID)",
                yaxis_title="Volume estimé (m³)",
                height=340,
                margin=dict(l=20, r=20, t=50, b=30),
                xaxis=dict(showgrid=True, gridcolor='#f1f5f9', showline=True, linecolor='#e2e8f0'),
                yaxis=dict(showgrid=True, gridcolor='#f1f5f9', showline=True, linecolor='#e2e8f0', tickformat='.2f')
            )
        except Exception as e:
            print(f"❌ Erreur graphique volume: {e}")
            fig_volume = create_empty_figure("Erreur graphique volume")
        
        return fig_priorite, fig_statut, fig_volume
        
    except Exception as e:
        print(f"❌ Erreur générale stats graphiques : {e}")
        fig_error = create_empty_figure(f"Erreur: {str(e)}")
        return fig_error, fig_error, fig_error

# ==================== CALLBACK NAVIGATION ====================
@app.callback(
    Output("page-content", "children"),
    Input("url", "pathname")
)
def display_page(pathname):
    if pathname == "/liste":
        return page_liste()
    elif pathname == "/stats":
        return page_stats()
    elif pathname == "/tester":
        return page_tester()
    else:
        return page_dashboard()

# ------------ Lancement ------------
if __name__ == "__main__":
    print("SANUYA Dashboard")
    print("http://localhost:8050")
    app.run(debug=True, dev_tools_ui=False, host='0.0.0.0', port=8050)